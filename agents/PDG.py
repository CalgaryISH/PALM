
# note to self:  this should be merged with PDG.py
import argparse

import json
import os
import re
from collections import Counter
from typing import Dict, Any, List, Tuple

import openpyxl


AES_LIST_ADDROUNDKEY = [
    "AddRoundKey","add_rnd_key", "ARKey", "add_round_key_module", "ARKeyModule", "add_rnd_key_fn", "AddRoundKeyTransform", "add_rkey","ARKeyTransform", "xor_add_key", "Add_RndKey_XOR", "RoundKeyAdder",
    "RoundKeyModule", "RndKeyTransform", "RndKey_XOR", "RoundKeyXOR_Module", "key_xor_rnd", "AddKeyXOR", "KeyXORAdder", "KeyXORModule", "AddRndXORKey", "AES_AddRndKey", "AddRoundKeyCore", "KeyXORCore", "AES_KeyXOR",
    "AES_RoundXOR","Add_Rnd_Key_AES", "RoundKey_XOR", "RndKeyAdder", "key_add_xor", "AESKeyAddModule", "AddRndKeyFunction", "ARKey_XOR","AESKeyXORModule", "AddKeyXORFn","XORKey_Round", "KeyAdd_Rnd", "round_key_combiner", "KeyXORAdderModule",
    "AES_RndKey_XOR", "RoundKeyXORFn", "RndKeyAdderAES", "RoundKeyCombiner", "AESRndKeyCombiner", "AESRndKeyAdder", "AES_Key_Adder","AddRndKeyTransform","RndKey_XOR_Fn", "AES_KeyRound", "AES_RKeyAdder", "KeyAdderAES",
     "AESKeyAdderModule", "KeyAdderRndXOR", "AESAddKeyXOR", "RndKeyXOR_Adder", "RndKey_Adder_XOR","AES_RKey_Combiner", "RoundKeyXORAdder", "AES_RndKey_Core", "AES_AddKey_XOR", "AddKey_RoundXOR", "AES_XOR_Key", "RoundKeyAdderAES",
    "KeyAdderXOR", "AddRndKey_Combiner" ,"RoundKeyXORModule", "AES_XOR_Key_Combiner", "AddRndKey_AesCore", "RoundKey_Add_XOR", "AES_KeyAdderFn", "AES_AddRndKeyFn", "RndKey_Add_Transform","KeyAdderXORModule" , "RoundXORKey", "AESRoundKeyFn",
    "XORKeyAdder" , "AESKeyTransform","KeyXorAddAES","AES_KeyXorAdder", "RoundKeyAES", "RoundXORCombiner" , "AddKeyTransformAES","RndKeyCombiner", "AESKeyXORCombiner", "KeyAddTransformAES", "XORKeyRoundAES", "RndKeyCoreAES", "AddXOR_RKey_AES",
    "AESKey_RndXOR_Adder","RoundKeyXORCore", "KeyAdderCoreAES", "AESRndKeyAdderXOR", "RndKeyAESAdder", "RoundXORModuleAES",
    "AddKeyRndModuleAES","XOR_RndKey_Adder", "AES_KeyAdderXORCore", "RoundKeyAdderCoreAES", "AESRndKeyAdderFn"
]


AES_LIST_SBOX = [
    "sbox", "S_box", "SBOX", "s_box_module", "sbox_unit", "sbox_8bit", "AES_sbox", "aes_S_box" , "s_box_aes", "sbox_transformation",
    "sbox_substitution", "substitution_box", "aes_sub_box", "sub_box",    "sbox_logic", "sbox_lookup", "sbox_lookup_table", "sbox_lut",
     "sbox_transform", "sbox_func", "sbox_core", "sbox_proc","aes_sbox_module", "sbox_function", "sbox_processor",
     "sbox_transformer", "sbox_mapper", "aes_sbox_processor" ,  "sbox_data_module", "sbox_engine", "sbox_unit_core", "sbox_handler",
     "sbox_data_proc", "sbox_op", "sbox_data_op", "aes_sbox_lut", "sbox_matrix", "sbox_sub_func", "sbox_8bit_transform",
    "aes_sbox_logic", "sbox_mapping", "sbox_table", "aes_sbox_unit", "sbox_processor_8bit", "sbox_module_8bit", "aes_sbox_transform",
    "aes_sbox_table", "sbox_byte_sub", "sbox_matrix_op" , "sbox_lookup_func", "sbox_transform_unit", "aes_sbox_lookup",
    "sbox_lut_module" , "sbox_data_transform", "aes_sbox_engine", "sbox_map_unit", "sbox_matrix_processor", "sbox_data_mapper",
    "aes_sbox_table_unit", "sbox_lookup_engine", "sbox_transform_core", "sbox_op_module", "sbox_mapping_proc", "aes_sbox_mapper",
    "sbox_byte_transformation", "sbox_matrix_logic", "sbox_func_unit", "sbox_byte_map", "aes_sbox_substitution", "sbox_core_func",
    "sbox_transformer_module", "sbox_proc_engine", "sbox_op_8bit", "aes_sbox_func", "aes_sbox_lut_unit","sbox_lut_processor",    "sbox_data_table","sbox_substitution_table", "sbox_lut_core", "aes_sbox_matrix" ,"sbox_lookup_engine_8bit", "sbox_processor_unit", "aes_sbox_byte_map","sbox_logic_module", "sbox_subs_func" ,"aes_sbox_byte_proc","sbox_lookup_processor", "sbox_matrix_table","sbox_byte_lookup","sbox_data_func", "sbox_data_engine","sbox_byte_substitute","sbox_processor_core", "aes_sbox_8bit_core","sbox_map_processor","sbox_sub_engine", "sbox_byte_table","sbox_data_handler", "sbox_table_unit", "sbox_sub_core"
]

AES_LIST_SHIFTROWS = [
     "shift_rows", "shiftrows", "ShiftRows", "SHIFTROWS", "shift_rows_func",   "shift_rows_module", "Shift_Rows", "Shift_Rows_Module", "shiftRows",
     "shift_rows_block", "ShiftRowsBlock", "SHIFT_ROWS_BLOCK", "shift_rows_logic", "ShiftRowsLogic", "SHIFT_ROWS_LOGIC", "sr_func", "sr_module", "SR_block",
    "sr_logic", "sr_transform" , "ShiftRowsTransform", "SHIFTROWS_TRANSFORM" , "sr_transform_block", "shift_rows_cyclic", "ShiftRowsCyclic", "sr_cyclic",
    "ShiftRowsMod", "shiftrows_cyclic_block", "ShiftRows_Unit", "SHIFTROWS_UNIT", "sr_cyclic_transform", "shift_rows_cycles", "ShiftRows_Cycle_Block",
    "SHIFTROWS_CYCLE", "Shift_Rows_Cyclic_Block","sr_transform_logic",  "sr_cyclic_logic", "shift_rows_matrix","ShiftRowsMatrix", "SHIFT_ROWS_MATRIX",    "sr_cycle_shift", "SR_shift_block", "shift_rows_design", "shift_rows_circuit",    "sr_circuit_logic","SR_logic_module", "sr_logic_unit", "shift_rows_control",
    "shiftrows_hdl", "ShiftRowsHDL", "sr_control_unit", "shiftrows_unit",  "ShiftRows_Circuits","ShiftRowsHDL_Module", "ShiftRowsUnit_Module",    "shiftrows_processing", "sr_processing_unit", "ShiftRowsProcessing",    "SHIFT_ROWS_CIRCUIT","shiftrows_hdl_block", "sr_hdl_logic", "sr_hdl_shift", "ShiftRows_Function","ShiftRows_Cycles", "shiftrows_process_unit",    "sr_hdl_module","shiftrows_transformer", "SHIFTROWS_CIRCUIT_LOGIC",    "sr_shift_hdl","shiftrows_hdl_unit", "shiftrows_fsm", "ShiftRowsFSM",
    "shiftrows_block_hdl","SHIFTROWS_FSM" ,"shiftrows_logic_block", "ShiftRows_Shift_Block" , "sr_fsm_block", "sr_shift_circuit",
    "shiftrows_logic_hdl","ShiftRows_Shift_Module", "sr_fsm_hdl",  "shiftrows_fsm_circuit","ShiftRows_Processing_Unit", "SHIFTROWS_Processing",
    "sr_shift_logic_hdl", "shiftrows_fsm_shift", "ShiftRows_FSM_Shift",  "sr_cycles_logic_hdl","ShiftRowsHDL_Shift", "ShiftRowsFSM_Logic"
]
AES_LIST_KEYEXPANSION = [
     'KEYEXPANSION', 'key_expansion', 'KeyExpansion', 'Key_Expander', 'Key_Expand_128', 'Key_Expand_192', 'Key_Expand_256',   'AES_Key_Expansion', 'AES_KeyGen', 'KeyGen_Module', 'KeyExpansion_Mod', 'Exp_Key', 'AES128_KeyExp', 'AES192_KeyExp',
    'AES256_KeyExp', 'RoundKey_Expander', 'RoundKey_Expand', 'KeyScheduler', 'KeySched_Mod', 'KeyGen_Core', 'KeyExp_Mod','KeyExpansionCore', 'AES_KeySched', 'KeyExpansionUnit', 'KeyExpansionBlock', 'KeyExpansionLogic','AES_KeyExpansion_Module',
    'RoundKeyGen_Module','KeyExpansionEngine', 'RoundKeyGenerator', 'KeyExpansionRoutine', 'KeyExpansionController','RoundKeyExp_Core', 'AES128_KeyExpansion_Core', 'KeyExpansionGenerator', 'KeyExpanderUnit', 'KeySchedulerEngine',
    'KeySchedule_Gen', 'AES_Key_Expand_Module', 'KeyExpansionTop' , 'AES_KeyGen_Module', 'AESKeyExp_Module', 'KeyGenUnit', 'KeySched_Controller', 'KeyScheduler_Block','AES_KeyExpUnit', 'KeyExpEngine', 'KeyExpansionProcessor', 'KeyExp_Module',
    'AESKeyExp_Mod','AESKeyExpansion', 'AES_KeyScheduler', 'AES_KeyExpansion_Core', 'AESKeySched_Logic', 'RoundKeySched_Module','KeyExpansion_Controller', 'AESKeyExpandBlock', 'RoundKey_ExpansionCore', 'KeySchedulerUnit','KeyExpansionTop_Module','AES_KeyExp_Core', 'AES128_KeySched_Core', 'KeyGen_Block', 'KeyExpBlock', 'AES_KeySchedulerUnit', 'KeyExpandProcessor',
    'KeySched_Module','AES_KeyExpand_Gen', 'RoundKey_Expander_Mod', 'KeyExpansion_Engine', 'AES_KeyExp_Controller','KeyExpTop_Module', 'KeyExpansionBlock_Core', 'KeyExp_Controller' ,'AES128_KeyExpansion_Unit','AES128KeyExp','AESKeySchedulerCore','AES_KeyExpand_Top', 'AES_KeyExp_Unit', 'KeyExpansion128_Core', 'AES_KeySched_Unit', 'AES_KeySchedulerBlock', 'KeyExpand_128_Block', 'KeyExpand128_Module', 'AESKeySchedulerBlock', 'AES128_KeyScheduler_Mod',
    'AES_KeyExp_Processor', 'KeyExp_Module128','AES128KeySched', 'AES_KeyExpansion_Unit', 'KeySched_Expander', 'AES_KeySched_Block','AES128_KeyExpansionBlock','AES128_KeyExp_Processor', 'AES_KeyExp_Gen_Block', 'KeySched_Top',
    'KeyExp_Gen_Block', 'KeyExpansionRoutine_Mod', 'KeySchedulerProcessor'
]
RSA_LIST_TOP = [
    "rsa_top", "rsa_core","rsa_engine", "rsa_main", "rsa_unit","rsa_top_module",    "rsa_top_level","rsa_wrapper", "rsa_controller", "rsa_processor", "modexp_top" , "modexp_core", "montgomery_core","montgomery_top","rsa_datapath","rsa_pipeline", "rsa_compute", "rsa_main_core", "rsa_block",
    "rsa_modexp", "rsa_montgomery", "exp_core", "exp_unit"
]

SHA_LIST_TOP = [
    "top_main", "sha_top","sha_core", "sha256_top", "sha256_core" , "sha2_core","sha512_top", "sha512_core", "sha_engine","sha_main", "sha_unit","sha_controller" , "sha_wrapper", "sha_top_level", "sha_datapath", "sha_pipeline","sha_round", "sha_sched", "message_schedule", "sha_transform"
]
FSM_LIST_TOP = [
    "fsm", "fsm_module", "state_machine", "ctrl_fsm", "control_fsm" , "controller_fsm","fsm_core", "fsm_top", "fsm_unit", "state_ctrl", "state_ctrl_fsm", "controller", "ctrl", "state", "next_state_logic","fsm_logic"
]

IP_REGISTRY: Dict[str, Dict[str, Dict[str, List[str]]]] = {
    "AES": {"functions": {
        "AddRoundKey":AES_LIST_ADDROUNDKEY,
        "SBox":AES_LIST_SBOX,
        "ShiftRows": AES_LIST_SHIFTROWS,
        "KeyExpansion":AES_LIST_KEYEXPANSION,
    }},
    "RSA": {"functions":{"top_module":RSA_LIST_TOP}},
    "SHA": {"functions": {"top_main":SHA_LIST_TOP}},
    "FSM": {"functions":{"fsm_module": FSM_LIST_TOP}},
}

def sorensen_dice_coefficient(str1: str, str2: str) -> float:
    s1 = str(str1)
    s2 =str(str2)
    b1 =set(s1[i:i + 2] for i in range(max(0, len(s1) - 1)))
    b2 =set(s2[i:i + 2] for i in range(max(0, len(s2) - 1)))
    tot =len(b1) + len(b2)
    if tot ==  0:
        return 0.0
    return 2 * len(b1 & b2) / tot

def _normalize_port_direction(raw: Any) -> str:
    if raw is None:
        return ""
    s= str(raw).strip().lower()
    if s in {"in", "input","portdirection.in","dir_in"}:
        return "input"
    if s in {"out", "output","portdirection.out", "dir_out"}:
        return "output"
    if s in { "inout", "portdirection.inout","ref","dir_inout"}:
        return "inout"
    if isinstance(raw, dict):
        k=str(raw.get("kind", "")).lower()
        if k == "in":
            return "input"
        if k =="out":
            return "output"
        if k == "inout":
            return "inout"
    return ""
def _port_direction_from_member(member: Dict[str, Any]) -> str:
    for key in ("direction","dir","portDirection","io"):
        if key in member:
            d = _normalize_port_direction(member.get(key))
            if d:
                    return d
    decl = member.get("decl", {})
    if isinstance(decl, dict):
        for key in ("direction","dir","portDirection","io"):
            if key in decl:
                d = _normalize_port_direction(decl.get(key))
                if d:
                        return d
    return ""



class PDGBuilder:
    def __init__(self, module_node=None):
            self.pdg: Dict[str, Dict[str, Any]] = {}
            self.input_list: List[str] = []
            self.visited= {}
            self.max_depth_cache = {}
            self.modules_dict: Dict[str, Any] = {}
            self.variables: List[Dict[str, Any]] = []
            self.current_module_name = ''
            self.module_instances = {}
            if module_node:
                self.build_pdg(module_node)

    def find_paths_to_inputs(self,start_node):
        paths = []
        stack = [(start_node, [start_node])]
        while stack:
            (node, path) = stack.pop()
            if node in self.input_list:
                 paths.append(path)
            else:
                for neighbor in self.pdg[node]['connections']:
                    if neighbor not in path:
                        stack.append((neighbor, path + [neighbor]))
        return paths

    def calculate_centroid(self):
        centroid_scores = {}
        for node_name, node_data in self.pdg.items():
            if node_data['type'] == 'Variable':
                paths= self.find_paths_to_inputs(node_name)
                total_weight = 0
                for path in paths:
                    path_weight = 0
                    for node in path:
                            operators= self.pdg[node].get('operators', {})
                            path_weight += sum(operators.values())
                    total_weight+= path_weight
                centroid_scores[node_name] = total_weight

        max_score = max(centroid_scores.values()) if centroid_scores else 0
        if max_score == 0:
            for node_name in centroid_scores:
                self.pdg[node_name]['centroid'] = 0
        else:
            for node_name, score in centroid_scores.items():
                self.pdg[node_name]['centroid'] = score / max_score

    def _append_variable(self, var_map: Dict[str, Dict[str, Any]], record: Dict[str, Any], is_port: bool):
        name = record['name']
        if name not in var_map:
            var_map[name] = record
            return
        existing = var_map[name]
        existing_is_port = existing.get('_kind') =='Port' and existing.get('type') in ('input', 'output', 'inout')
        if is_port:
            var_map[name] = record
        else:
            if existing_is_port:
                return

            if existing.get('type') in ('Unknown', '', None, 'logic') and record.get('type') not in ('Unknown', '', None):
                var_map[name]= record

    def extract_variables_and_statements_from_module(self, node):
        variables_map: Dict[str, Dict[str, Any]] = {}
        statements = []
        instances = []
        if not node:
            return [], statements, instances

        members = node.get('members', [])
        for member in members:
            kind = member.get('kind', '')

            if kind in ['Port', 'Variable', 'Net', 'Parameter']:
                var_name = member.get('name')
                var_type = ""
                length = 1

                port_dir = _port_direction_from_member(member) if kind == 'Port' else ""
                if port_dir:
                    var_type = port_dir  #input |output |inout

                #bit width
                type_info = member.get('type', '')
                if isinstance(type_info, dict):
                    packed = type_info.get("packedRange")
                    if isinstance(packed, dict) and "left" in packed and "right" in packed:
                        try:
                            length = abs(int(packed["left"]) - int(packed["right"])) + 1
                        except Exception:
                            length = 1
                    if not var_type:
                        base = type_info.get("base", "") or type_info.get("kind", "")
                        var_type = str(base or "logic").lower()
                else:
                    pattern = r'(\w+)(\[(\d+):(\d+)\])?'
                    match = re.match(pattern, str(type_info))
                    if match:
                        base_type = match.group(1)
                        if not var_type:
                            if kind == 'Net':
                                net_type = member.get('netType', {}).get('name', 'wire')
                                var_type = net_type if net_type else base_type
                            else:
                                var_type = base_type
                        if match.group(3) and match.group(4):
                            index1 = int(match.group(3)); index2 = int(match.group(4))
                            length = abs(index1 - index2) + 1
                        else:
                            length = 1
                    else:
                        if not var_type:
                            var_type = 'unknown'
                        length = 1

                rec = {'name': var_name, 'type': var_type, 'length': length, '_kind': kind}
                self._append_variable(variables_map, rec, is_port=(kind == 'Port'))

                if 'initializer' in member:
                    assignment = {
                        'kind': 'VariableInitialization',
                        'left': {'kind': 'NamedValue', 'symbol': f"{member.get('addr', '')} {var_name}"},
                        'right': member['initializer']
                    }
                    statements.append(assignment)

            elif kind in ['ContinuousAssign']:
                statements.append(member)

            elif kind in ['ProceduralBlock', 'Always', 'Initial']:
                statements.append(member)
                self.extract_variables_from_statement(member, list(variables_map.values()))

            elif kind == 'Instance':
                instances.append(member)

        variables = []
        for v in variables_map.values():
                v.pop('_kind', None)
                variables.append(v)

        return variables, statements, instances

    def extract_variables_from_statement(self, stmt, variables):
        if not isinstance(stmt, dict):
            return
        kind = stmt.get('kind', '')
        if kind in ['ProceduralBlock', 'Block', 'Timed', 'List', 'StatementBlock', 'Conditional', 'ForLoop']:
            body = stmt.get('body') or stmt.get('list') or stmt.get('stmts') or []
            if isinstance(body, dict):
                self.extract_variables_from_statement(body, variables)
            elif isinstance(body, list):
                for s in body:
                    self.extract_variables_from_statement(s, variables)

        elif kind == 'VariableDeclaration':
            var_name = stmt.get('symbol', '').split(' ')[-1]
            if var_name and not any(var['name'] == var_name for var in variables):
                var_type = stmt.get('type', 'Unknown')
                variables.append({'name': var_name, 'type': var_type, 'length': 'Unknown'})

        elif kind == 'ExpressionStatement':
            expr = stmt.get('expr', {})
            if expr.get('kind') == 'Assignment':
                left = expr.get('left', {})
                var_name = self.extract_target_name(left)
                if var_name and not any(var['name'] == var_name for var in variables):
                    variables.append({'name': var_name, 'type': 'Unknown', 'length': 'Unknown'})

        elif kind == 'Conditional':
            ifTrue = stmt.get('ifTrue', {})
            ifFalse = stmt.get('ifFalse', {})
            self.extract_variables_from_statement(ifTrue, variables)
            self.extract_variables_from_statement(ifFalse, variables)

        elif kind == 'Case':
            items = stmt.get('items', [])
            for item in items:
                self.extract_variables_from_statement(item.get('stmt', {}), variables)
            default_case = stmt.get('defaultCase')
            if default_case:
                self.extract_variables_from_statement(default_case, variables)

    def extract_sources(self, expression):
        sources = {'names': [], 'ops': []}

        if isinstance(expression, list):
            for expr in expression:
                nested_sources = self.extract_sources(expr)
                sources['names'].extend(nested_sources['names'])
                sources['ops'].extend(nested_sources['ops'])
            return sources

        if not isinstance(expression, dict):
            return sources

        kind = expression.get('kind', '')
        if kind == 'NamedValue':
            if 'symbol' in expression:
                sources['names'].append(expression['symbol'].split(' ')[-1])

        elif kind == 'BinaryOp':
            operator = expression.get('op') or expression.get('operator')
            if operator:
                sources['ops'].append(operator)
            left_sources = self.extract_sources(expression.get('left', {}))
            right_sources = self.extract_sources(expression.get('right', {}))
            #sources['names'].extend( )

            sources['names'].extend(left_sources['names'] + right_sources['names'])
            sources['ops'].extend(left_sources['ops'] + right_sources['ops'])

        elif kind == 'UnaryOp':
            operator = expression.get('op') or expression.get('operator')
            if operator:
                sources['ops'].append(operator)
            operand_sources = self.extract_sources(expression.get('operand', {}))
            sources['names'].extend(operand_sources['names'])
            sources['ops'].extend(operand_sources['ops'])

        elif kind == 'ConditionalOp':
                cond_sources = self.extract_sources(expression.get('cond', {}))
                true_sources = self.extract_sources(expression.get('left', {}))
                false_sources = self.extract_sources(expression.get('right', {}))
                sources['names'].extend(cond_sources['names'] + true_sources['names'] + false_sources['names'])
                sources['ops'].extend(cond_sources['ops'] + true_sources['ops'] + false_sources['ops'])

        elif kind == 'ElementSelect':
            value_sources = self.extract_sources(expression.get('value', {}))
            selector_sources = self.extract_sources(expression.get('selector', {}))
            sources['names'].extend(value_sources['names'] + selector_sources['names'])

        elif kind == 'Conversion':
            operand_sources = self.extract_sources(expression.get('operand', {}))
            sources['names'].extend(operand_sources['names'])
            sources['ops'].extend(operand_sources['ops'])

        elif kind == 'IntegerLiteral':
            pass

        return sources

    def extract_condition_str(self, expr):
        kind = expr.get('kind', '')
        if kind == 'NamedValue':
            return expr['symbol'].split(' ')[-1]
        elif kind == 'BinaryOp':
            left=self.extract_condition_str(expr['left'])
            right=self.extract_condition_str(expr['right'])
            op = expr['op']
            return f'({left} {op} {right})'
        elif kind == 'UnaryOp':
            operand = self.extract_condition_str(expr['operand'])
            op = expr['op']
            return f'({op}{operand})'
        elif kind == 'IntegerLiteral':
            return expr.get('value', '')
        else:
            return ''

    def extract_target_name(self, expr):
        kind = expr.get('kind', '')
        if kind in ['NamedValue', 'Variable']:
            if 'symbol' in expr:
                return expr['symbol'].split(' ')[-1]
        elif kind in ['ElementSelect', 'RangeSelect']:
            if 'value' in expr:
                return self.extract_target_name(expr['value'])
        elif kind == 'MemberAccess':
            if 'parent' in expr and 'member' in expr:
                base_name = self.extract_target_name(expr['parent'])
                member_name = expr['member']
                return f"{base_name}.{member_name}"
        return None

    def ensure_nodes_in_pdg(self, nodes):
        for node in nodes:
            if node not in self.pdg:
                self.pdg[node] = {
                    'name': node,
                    'type': 'Variable',
                    'connections': [],
                    'conditions': [],
                    'operators': Counter(),
                    'details': {},
                }

    def update_pdg(self, target, sources, conditions, operators=None):
        if target is None:
            return
        if target not in self.pdg:
            self.pdg[target] = {
                'name': target,
                'type': 'Variable',
                'connections': [],
                'conditions': [],
                'operators': Counter(),
                'details': {},
            }
        self.ensure_nodes_in_pdg(sources)
        self.pdg[target]['connections'].extend(sources)
        self.pdg[target]['connections'] = sorted(set(self.pdg[target]['connections']))
        self.pdg[target]['conditions'].extend(conditions)
        self.pdg[target]['conditions'] = sorted(set(self.pdg[target]['conditions']))
        if operators is not None:
            self.pdg[target]['operators'].update(operators)

    def add_connections_conditional(self, node, current_conditions, parent_nodes):
        if node.get('kind') != 'Conditional':
            self.process_statement(node, current_conditions, parent_nodes)
            return

        conditions = node.get('conditions', [])
        for condition in conditions:
            condition_expr = condition.get('expr', {})
            condition_str = self.extract_condition_str(condition_expr)

            cond_sources = self.extract_sources(condition_expr)
            cond_node_name = f"cond_{id(condition)}"
            self.ensure_nodes_in_pdg(cond_sources['names'])
            self.ensure_nodes_in_pdg(parent_nodes)

            self.pdg[cond_node_name] = {
                'name': cond_node_name,
                'type': 'Condition',
                'connections': list(set(cond_sources['names'] + parent_nodes)),
                'conditions': [],
                'operators': Counter(),
                'details': {},
            }

            ifTrue = node.get('ifTrue', {})
            next_parents = parent_nodes + [cond_node_name]
            if ifTrue:
                self.process_statement(ifTrue, current_conditions + [condition_str], next_parents)

            ifFalse = node.get('ifFalse', {})
            if ifFalse:
                condition_negation_str = f'!({condition_str})'
                self.process_statement(ifFalse, current_conditions + [condition_negation_str], next_parents)

    def add_connections_case(self, case_expr, current_conditions, parent_nodes):
        case_cond = case_expr.get('expr', {})
        case_expr_str = self.extract_condition_str(case_cond)
        case_sources = self.extract_sources(case_cond)
        case_node_name = f"case_{id(case_expr)}"
        self.ensure_nodes_in_pdg(case_sources['names'])
        self.ensure_nodes_in_pdg(parent_nodes)

        self.pdg[case_node_name] = {
            'name': case_node_name,
            'type': 'Condition',
            'connections': case_sources['names'] + parent_nodes,
            'conditions': [],
            'operators': Counter(),
            'details': {},
        }

        items = case_expr.get('items', [])
        for item in items:
            item_exprs = item.get('expressions', [])
            for expr in item_exprs:
                expr_str = self.extract_condition_str(expr)
                condition_str = f'({case_expr_str} == {expr_str})'
                self.process_statement(item.get('stmt', {}), current_conditions + [condition_str],
                                       parent_nodes + [case_node_name])

        default_case = case_expr.get('defaultCase')
        if default_case:
            condition_str = f'default({case_expr_str})'
            self.process_statement(default_case, current_conditions + [condition_str],
                                   parent_nodes + [case_node_name])

    def process_statement(self, node, current_conditions, parent_nodes=None):
        if parent_nodes is None:
            parent_nodes = []
        kind = node.get('kind', '')

        if kind == 'ExpressionStatement':
            expr = node.get('expr', {})
            if 'left' in expr and 'right' in expr:
                target = self.extract_target_name(expr['left'])
                sources = self.extract_sources(expr['right'])
                operators = Counter(sources['ops'])
                connections = sources['names'] + parent_nodes
                self.update_pdg(target, connections, current_conditions, operators)

                if 'symbol' in expr.get('left', {}):
                    target2 = expr['left']['symbol'].split(' ')[-1]
                    self.update_pdg(target2, connections, current_conditions, operators)

        elif kind == 'ContinuousAssign':
            assignment = node.get('assignment', {})
            if 'left' in assignment and 'right' in assignment:
                target = self.extract_target_name(assignment['left'])
                sources = self.extract_sources(assignment['right'])
                operators = Counter(sources['ops'])
                connections = sources['names'] + parent_nodes
                self.update_pdg(target, connections, current_conditions, operators)

        elif kind == 'Block':
            body = node.get('body')
            if isinstance(body, dict):
                self.process_statement(body , current_conditions , parent_nodes)
            elif isinstance(body, list):
                for stmt in body:
                    self.process_statement(stmt, current_conditions, parent_nodes)

        elif kind == 'List':
            for stmt in node.get('list', []):
                self.process_statement(stmt , current_conditions, parent_nodes)

        elif kind == 'Conditional':
            self.add_connections_conditional(node, current_conditions, parent_nodes)

        elif kind == 'Case':
            self.add_connections_case(node , current_conditions , parent_nodes)

        elif kind == 'Timed':
            if 'stmt' in node:
                self.process_statement(node['stmt'], current_conditions, parent_nodes)

        elif kind == 'ProceduralBlock':
            self.process_statement(node.get('body', {}), current_conditions, parent_nodes)

        elif kind == 'Assignment':
            left = node.get('left', {})
            target = left.get('symbol', '').split(' ')[-1] if 'symbol' in left else None
            sources = self.extract_sources(node.get('right', {}))
            connections = sources['names'] + parent_nodes
            self.update_pdg(target, connections, current_conditions)

        elif kind == 'ForLoop':
            initializer = node.get('init')
            condition = node.get('cond')
            increment = node.get('inc')
            body = node.get('body')

            if initializer:
                self.process_statement(initializer, current_conditions, parent_nodes)
            if condition:
                _ = self.extract_sources(condition)
            if increment:
                _ = self.extract_sources(increment)
            if body:
                if isinstance(body, list):
                    for stmt in body:
                        self.process_statement(stmt, current_conditions, parent_nodes)
                elif isinstance(body, dict):
                    self.process_statement(body, current_conditions, parent_nodes)

    def handle_node_types(self, node):
        kind = node.get('kind', '')
        if kind in ['Port', 'Variable', 'Net', 'Parameter']:
            name = node.get('name', '').split(' ')[-1]
            if kind == 'Port':
                direction = _port_direction_from_member(node)
                if direction in ('input', 'inout'):
                    self.input_list.append(name)
            if name and (name not in self.pdg):
                self.pdg[name] = {
                    'name': name,
                    'type': 'Variable',
                    'connections': [],
                    'conditions': [],
                    'operators': Counter(),
                    'details': node,
                }

        elif kind == 'ContinuousAssign':
            assignment = node.get('assignment', {})
            if 'left' in assignment and 'right' in assignment:
                target = self.extract_target_name(assignment['left'])
                sources = self.extract_sources(assignment['right'])
                self.ensure_nodes_in_pdg(sources['names'])
                if target:
                    self.update_pdg(target, sources['names'], [])

        elif kind == 'Instance':
            pass

        if 'body' in node and kind not in ['ProceduralBlock', 'Block', 'Conditional', 'List', 'Timed']:
            if isinstance(node['body'], dict):
                self.handle_node_types(node['body'])
            elif isinstance(node['body'], list):
                for item in node['body']:
                    self.handle_node_types(item)
        if 'members' in node:
            for member in node['members']:
                self.handle_node_types(member)
        if 'stmts' in node:
            for stmt in node['stmts']:
                self.handle_node_types(stmt)
        if 'expr' in node and node['expr'].get('kind', '') == 'ConditionalOp':
            self.process_statement(node, [])

    def calculate_dependence_depth(self, start_name):
        def dfs(name):
            if name in self.max_depth_cache:
                return self.max_depth_cache[name]
            if name in self.visited:
                return 0
            self.visited.add(name)
            max_depth = 0
            if name in self.pdg:
                node = self.pdg[name]
                connections = node['connections']
                conditions = node['conditions']
                count_cnd = sum(1 for item in connections if str(item).startswith('cond_'))
                count_cse = sum(1 for item in connections if str(item).startswith('case_'))

                if len(conditions) >= 1:
                    for cond in conditions:
                        cond_depth = dfs(cond)
                        if cond_depth > max_depth:
                            max_depth = cond_depth
                    if len(connections) >= 1:
                        for conn in connections:
                                if str(conn).startswith('cond_'):
                                    conn_depth = 0
                                else:
                                    conn_depth = dfs(conn)
                                if conn_depth > max_depth:
                                    max_depth= conn_depth
                else:
                    for conn in connections:
                        if str(conn).startswith('cond_'):
                            max_depth = max(max_depth, 0)
                        else:
                            conn_depth = dfs(conn)
                            if conn_depth > max_depth:
                                max_depth = conn_depth

                max_depth = count_cnd + max_depth + count_cse

            self.visited.remove(name)
            depth = max_depth + 1
            self.max_depth_cache[name] = depth
            return depth

        self.visited = set()
        self.max_depth_cache = {}
        total_depth = dfs(start_name)
        return total_depth - 1  #inputs depth 0

    def build_pdg(self, module_node):
        variables, statements, instances =  self.extract_variables_and_statements_from_module(module_node)
        self.variables = variables
        for var in variables:
            name = var['name']
            var_type = var['type']
            if var_type in ('input', 'inout'):
                self.input_list.append(name)
            if name not in self.pdg:
                self.pdg[name] = {
                    'name': name,
                    'type': 'Variable',
                    'connections': [],
                    'conditions': [],
                    'operators': Counter(),
                    'details': var,
                }

        for stmt in statements:
            self.process_statement(stmt, [])

        for instance in instances:
            self.handle_node_types(instance)

        self.calculate_centroid()

    def aggregate_operators(self):
        for node_name in list(self.pdg.keys()):
            self._visited_agg= set()
            operators= self.get_aggregated_operators(node_name)
            self.pdg[node_name]['aggregated_operators']= operators

    def get_aggregated_operators(self, node_name):
        if 'aggregated_operators' in self.pdg[node_name]:
            return self.pdg[node_name]['aggregated_operators']
        if not hasattr(self, '_visited_agg'):
            self._visited_agg = set()
        if node_name in self._visited_agg:
            return Counter()
        self._visited_agg.add(node_name)

        operators = Counter(self.pdg[node_name]['operators'])
        for conn in self.pdg[node_name]['connections']:
            if conn in self.pdg:
                conn_operators = self.get_aggregated_operators(conn)
                operators.update(conn_operators)
        self.pdg[node_name]['aggregated_operators'] = operators
        return operators

    def main(self, json_data, module_name):
        def build_modules_dict(node, out: Dict[str, Any]):
            if isinstance(node, dict):
                if node.get('kind') == 'InstanceBody':
                    mname= node.get('name')
                    if mname:
                        out[mname] = node
                    for member in node.get('members', []):
                        build_modules_dict(member, out)
                else:
                    for _, value in node.items():
                        if isinstance(value, (dict, list)):
                            build_modules_dict(value, out)
            elif isinstance(node, list):
                for item in node:
                    build_modules_dict(item,out)

        self.modules_dict = {}
        build_modules_dict(json_data, self.modules_dict)
        module_node = self.modules_dict.get(module_name)
        if module_node is None:
            print(f"module '{module_name}' not found in the JSON data.")
            return

        self.current_module_name = module_name
        self.build_pdg(module_node)
        self.aggregate_operators()

        for name, node in list(self.pdg.items()):
            if node['type'] in ('Variable', 'Condition'):
                self.visited = {}
                self.max_depth_cache = {}
                depth = self.calculate_dependence_depth(name)
                self.pdg[name]['depth'] = depth

    def get_pdg_depths(self):
        pdg_depths = {}
        for name, node in self.pdg.items():
            if node['type'] in ('Variable','Condition'):
                pdg_depths[name] = node.get('depth', 'N/A')
        return pdg_depths

    def get_variables(self):
        return self.variables


#######################


def build_modules_dict(node: Any, modules_dict: Dict[str, Any]) -> None:
    if isinstance(node, dict):
        if node.get("kind") == "InstanceBody":
            module_name= node.get("name")
            if module_name:
                modules_dict[module_name] = node
            for member in node.get("members", []):
                build_modules_dict(member,  modules_dict)
        else:
            for _, value in node.items():
                if isinstance(value, (dict, list)):
                    build_modules_dict(value, modules_dict)
    elif isinstance(node, list):
        for item in node:
            build_modules_dict(item, modules_dict)


def find_best_matching_module(modules_dict:Dict[str, Any], alias_list: List[str]) -> Tuple[str, Any]:
    best_name = None
    best_score = -1.0
    for alias in alias_list:
        al= alias.lower()
        for mod_name in modules_dict.keys():
            sc = sorensen_dice_coefficient(al, str(mod_name).lower())
            if sc > best_score:
                best_score = sc
                best_name = mod_name
    return best_name, modules_dict.get(best_name)

def main():
    parser = argparse.ArgumentParser(description="PDG batch analyzer (AES/RSA/SHA/FSM)")
    parser.add_argument("--family", required=True, choices=list(IP_REGISTRY.keys()),
                        help="IP family to use (AES, RSA, SHA, FSM)")
    parser.add_argument("--design_folder", required=True,
                        help="Folder containing *.json AST files")
    parser.add_argument("--summary", default=None,
                        help="Optional output filename for the summary workbook (default: Summary_<family>.xlsx)")
    args = parser.parse_args()

    family = args.family
    functions = IP_REGISTRY[family]["functions"]
    design_folder = args.design_folder

    if not os.path.isdir(design_folder):
        raise FileNotFoundError(f"Design folder not found: {design_folder}")
    design_files = [f for f in os.listdir(design_folder) if f.endswith(".json")]
    if not design_files:
        print(f"no *.json files found in {design_folder}")
        return

    summary_data: Dict[str, Dict[str, str]] = {}

    for design_file in design_files:
        json_file_path = os.path.join(design_folder, design_file)
        try:
            with open(json_file_path, "r", encoding="utf-8") as f:
                json_data =  json.load(f)
        except Exception as e:
            print(f("failed to read '{json_file_path}': {e}"))
            continue

        modules_dict: Dict[str, Any] = {}
        build_modules_dict(json_data, modules_dict)
        module_names= list(modules_dict.keys())

        excel_file = f'{os.path.splitext(design_file)[0]}.xlsx'
        workbook = openpyxl.Workbook()
        created_any_sheet = False
        summary_data[design_file] = {}

        for func_name, alias_list in functions.items():
            print(f"\nProcessing list: {func_name} | file: {design_file}")
            similarity_counter = Counter()

            for alias in alias_list:
                highest_similarity = 0.0
                most_similar_name =None
                for module_name in module_names:
                    similarity = sorensen_dice_coefficient(alias.lower() , module_name.lower())
                    if similarity > highest_similarity:
                        highest_similarity =similarity
                        most_similar_name = module_name
                if most_similar_name:
                    similarity_counter[most_similar_name]+= 1

            if similarity_counter:
                best_module, count = similarity_counter.most_common(1)[0]
                print(f" most repeated similar module in {func_name} : '{best_module}' (x{count})")
                summary_data[design_file][func_name] = best_module

                module_node = modules_dict.get(best_module)
                if module_node is None:
                    print(f"module '{best_module}'not found in design. Skipping sheet '{func_name}'.")
                    continue
                builder = PDGBuilder(module_node)
                builder.main(json_data , best_module)

                pdg_depths = builder.get_pdg_depths()
                variables = builder.get_variables()
                variable_info = {v['name']: v for v in variables}
                sheet_title = func_name[:31]
                if sheet_title in workbook.sheetnames:
                    sheet = workbook[sheet_title]
                else:
                    sheet = workbook.create_sheet(title=sheet_title)
                if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) > 1:
                        try:
                            workbook.remove(workbook['Sheet'])
                        except Exception:
                            pass

                sheet.append(['Variable Name', 'Type','Bit Width', 'PDG_Depth','Num_Operators','Centroid'])

                for var_name, depth in pdg_depths.items():
                    info = variable_info.get(var_name, {'type': 'Unknown', 'length': 'Unknown'})
                    pdg_node = builder.pdg.get(var_name, {})
                    num_ops = sum(pdg_node.get('aggregated_operators', {}).values()) / 2 if pdg_node else 0
                    centroid = pdg_node.get('centroid', 0) if pdg_node else 0
                    sheet.append([var_name, info.get('type', 'Unknown'), info.get('length', 'Unknown'),
                                  depth, num_ops, centroid])

                created_any_sheet = True
            else:
                print(f"no similar modules found for {func_name} in {design_file}")
                summary_data[design_file][func_name] = 'None'

        if not created_any_sheet:
            sheet = workbook.active
            sheet.title = "Info"
            sheet.append(["no function sheets were created for this design.."])

        try:
            workbook.save(excel_file)
            print(f"wrote: {excel_file}")
        except Exception as e:
            print(f"failed to save '{excel_file}': {e}")

    summary_excel_file = args.summary or f"Summary_{family}.xlsx"
    summary_wb = openpyxl.Workbook()
    summary_sheet = summary_wb.active
    summary_sheet.title = "Summary"
    headers = ['Design File'] + list(functions.keys())
    summary_sheet.append(headers)

    for design_file, funcs in summary_data.items():
        row = [design_file]
        for func_name in functions.keys():
            row.append(funcs.get(func_name, 'None'))
        summary_sheet.append(row)

    try:
        summary_wb.save(summary_excel_file)
        print(f"\n  summary saved to:::::::  {summary_excel_file}")
    except Exception as e:
        print(f"failed to save summary '{summary_excel_file}': {e}")

if __name__== "__main__":

    main()
