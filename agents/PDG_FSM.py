#note to self:  this should be merged with PDG.py


import argparse
import json
import os
import re
from collections import Counter, defaultdict
from typing import Dict, Any, List, Tuple, Optional

import openpyxl
AES_LIST_ADDROUNDKEY = [
    "AddRoundKey","add_rnd_key", "ARKey", "add_round_key_module", "ARKeyModule", "add_rnd_key_fn", "AddRoundKeyTransform", "add_rkey","ARKeyTransform", "xor_add_key", "Add_RndKey_XOR", "RoundKeyAdder",
    "RoundKeyModule", "RndKeyTransform", "RndKey_XOR", "RoundKeyXOR_Module", "key_xor_rnd", "AddKeyXOR", "KeyXORAdder", "KeyXORModule", "AddRndXORKey", "AES_AddRndKey", "AddRoundKeyCore", "KeyXORCore", "AES_KeyXOR",
    "AES_RoundXOR","Add_Rnd_Key_AES", "RoundKey_XOR", "RndKeyAdder", "key_add_xor", "AESKeyAddModule", "AddRndKeyFunction", "ARKey_XOR","AESKeyXORModule", "AddKeyXORFn","XORKey_Round", "KeyAdd_Rnd", "round_key_combiner", "KeyXORAdderModule",
    "AES_RndKey_XOR", "RoundKeyXORFn", "RndKeyAdderAES", "RoundKeyCombiner", "AESRndKeyCombiner", "AESRndKeyAdder", "AES_Key_Adder","AddRndKeyTransform","RndKey_XOR_Fn", "AES_KeyRound", "AES_RKeyAdder", "KeyAdderAES",
     "AESKeyAdderModule", "KeyAdderRndXOR", "AESAddKeyXOR", "RndKeyXOR_Adder", "RndKey_Adder_XOR","AES_RKey_Combiner", "RoundKeyXORAdder", "AES_RndKey_Core", "AES_AddKey_XOR", "AddKey_RoundXOR", "AES_XOR_Key", "RoundKeyAdderAES",
    "KeyAdderXOR", "AddRndKey_Combiner" ,"RoundKeyXORModule", "AES_XOR_Key_Combiner", "AddRndKey_AesCore", "RoundKey_Add_XOR", "AES_KeyAdderFn", "AES_AddRndKeyFn", "RndKey_Add_Transform","KeyAdderXORModule" , "RoundXORKey", "AESRoundKeyFn",
    "XORKeyAdder" , "AESKeyTransform","KeyXorAddAES","AES_KeyXorAdder", "RoundKeyAES", "RoundXORCombiner" , "AddKeyTransformAES","RndKeyCombiner", "AESKeyXORCombiner", "KeyAddTransformAES", "XORKeyRoundAES", "RndKeyCoreAES", "AddXOR_RKey_AES",
    "AESKey_RndXOR_Adder","RoundKeyXORCore", "KeyAdderCoreAES", "AESRndKeyAdderXOR", "RndKeyAESAdder", "RoundXORModuleAES",
    "AddKeyRndModuleAES","XOR_RndKey_Adder", "AES_KeyAdderXORCore", "RoundKeyAdderCoreAES", "AESRndKeyAdderFn"
]


AES_LIST_SBOX = [
    "sbox", "S_box", "SBOX", "s_box_module", "sbox_unit", "sbox_8bit", "AES_sbox", "aes_S_box" , "s_box_aes", "sbox_transformation",
    "sbox_substitution", "substitution_box", "aes_sub_box", "sub_box",    "sbox_logic", "sbox_lookup", "sbox_lookup_table", "sbox_lut",
     "sbox_transform", "sbox_func", "sbox_core", "sbox_proc","aes_sbox_module", "sbox_function", "sbox_processor",
     "sbox_transformer", "sbox_mapper", "aes_sbox_processor" ,  "sbox_data_module", "sbox_engine", "sbox_unit_core", "sbox_handler",
     "sbox_data_proc", "sbox_op", "sbox_data_op", "aes_sbox_lut", "sbox_matrix", "sbox_sub_func", "sbox_8bit_transform",
    "aes_sbox_logic", "sbox_mapping", "sbox_table", "aes_sbox_unit", "sbox_processor_8bit", "sbox_module_8bit", "aes_sbox_transform",
    "aes_sbox_table", "sbox_byte_sub", "sbox_matrix_op" , "sbox_lookup_func", "sbox_transform_unit", "aes_sbox_lookup",
    "sbox_lut_module" , "sbox_data_transform", "aes_sbox_engine", "sbox_map_unit", "sbox_matrix_processor", "sbox_data_mapper",
    "aes_sbox_table_unit", "sbox_lookup_engine", "sbox_transform_core", "sbox_op_module", "sbox_mapping_proc", "aes_sbox_mapper",
    "sbox_byte_transformation", "sbox_matrix_logic", "sbox_func_unit", "sbox_byte_map", "aes_sbox_substitution", "sbox_core_func",
    "sbox_transformer_module", "sbox_proc_engine", "sbox_op_8bit", "aes_sbox_func", "aes_sbox_lut_unit","sbox_lut_processor",    "sbox_data_table","sbox_substitution_table", "sbox_lut_core", "aes_sbox_matrix" ,"sbox_lookup_engine_8bit", "sbox_processor_unit", "aes_sbox_byte_map","sbox_logic_module", "sbox_subs_func" ,"aes_sbox_byte_proc","sbox_lookup_processor", "sbox_matrix_table","sbox_byte_lookup","sbox_data_func", "sbox_data_engine","sbox_byte_substitute","sbox_processor_core", "aes_sbox_8bit_core","sbox_map_processor","sbox_sub_engine", "sbox_byte_table","sbox_data_handler", "sbox_table_unit", "sbox_sub_core"
]

AES_LIST_SHIFTROWS = [
     "shift_rows", "shiftrows", "ShiftRows", "SHIFTROWS", "shift_rows_func",   "shift_rows_module", "Shift_Rows", "Shift_Rows_Module", "shiftRows",
     "shift_rows_block", "ShiftRowsBlock", "SHIFT_ROWS_BLOCK", "shift_rows_logic", "ShiftRowsLogic", "SHIFT_ROWS_LOGIC", "sr_func", "sr_module", "SR_block",
    "sr_logic", "sr_transform" , "ShiftRowsTransform", "SHIFTROWS_TRANSFORM" , "sr_transform_block", "shift_rows_cyclic", "ShiftRowsCyclic", "sr_cyclic",
    "ShiftRowsMod", "shiftrows_cyclic_block", "ShiftRows_Unit", "SHIFTROWS_UNIT", "sr_cyclic_transform", "shift_rows_cycles", "ShiftRows_Cycle_Block",
    "SHIFTROWS_CYCLE", "Shift_Rows_Cyclic_Block","sr_transform_logic",  "sr_cyclic_logic", "shift_rows_matrix","ShiftRowsMatrix", "SHIFT_ROWS_MATRIX",    "sr_cycle_shift", "SR_shift_block", "shift_rows_design", "shift_rows_circuit",    "sr_circuit_logic","SR_logic_module", "sr_logic_unit", "shift_rows_control",
    "shiftrows_hdl", "ShiftRowsHDL", "sr_control_unit", "shiftrows_unit",  "ShiftRows_Circuits","ShiftRowsHDL_Module", "ShiftRowsUnit_Module",    "shiftrows_processing", "sr_processing_unit", "ShiftRowsProcessing",    "SHIFT_ROWS_CIRCUIT","shiftrows_hdl_block", "sr_hdl_logic", "sr_hdl_shift", "ShiftRows_Function","ShiftRows_Cycles", "shiftrows_process_unit",    "sr_hdl_module","shiftrows_transformer", "SHIFTROWS_CIRCUIT_LOGIC",    "sr_shift_hdl","shiftrows_hdl_unit", "shiftrows_fsm", "ShiftRowsFSM",
    "shiftrows_block_hdl","SHIFTROWS_FSM" ,"shiftrows_logic_block", "ShiftRows_Shift_Block" , "sr_fsm_block", "sr_shift_circuit",
    "shiftrows_logic_hdl","ShiftRows_Shift_Module", "sr_fsm_hdl",  "shiftrows_fsm_circuit","ShiftRows_Processing_Unit", "SHIFTROWS_Processing",
    "sr_shift_logic_hdl", "shiftrows_fsm_shift", "ShiftRows_FSM_Shift",  "sr_cycles_logic_hdl","ShiftRowsHDL_Shift", "ShiftRowsFSM_Logic"
]
AES_LIST_KEYEXPANSION = [
     'KEYEXPANSION', 'key_expansion', 'KeyExpansion', 'Key_Expander', 'Key_Expand_128', 'Key_Expand_192', 'Key_Expand_256',   'AES_Key_Expansion', 'AES_KeyGen', 'KeyGen_Module', 'KeyExpansion_Mod', 'Exp_Key', 'AES128_KeyExp', 'AES192_KeyExp',
    'AES256_KeyExp', 'RoundKey_Expander', 'RoundKey_Expand', 'KeyScheduler', 'KeySched_Mod', 'KeyGen_Core', 'KeyExp_Mod','KeyExpansionCore', 'AES_KeySched', 'KeyExpansionUnit', 'KeyExpansionBlock', 'KeyExpansionLogic','AES_KeyExpansion_Module',
    'RoundKeyGen_Module','KeyExpansionEngine', 'RoundKeyGenerator', 'KeyExpansionRoutine', 'KeyExpansionController','RoundKeyExp_Core', 'AES128_KeyExpansion_Core', 'KeyExpansionGenerator', 'KeyExpanderUnit', 'KeySchedulerEngine',
    'KeySchedule_Gen', 'AES_Key_Expand_Module', 'KeyExpansionTop' , 'AES_KeyGen_Module', 'AESKeyExp_Module', 'KeyGenUnit', 'KeySched_Controller', 'KeyScheduler_Block','AES_KeyExpUnit', 'KeyExpEngine', 'KeyExpansionProcessor', 'KeyExp_Module',
    'AESKeyExp_Mod','AESKeyExpansion', 'AES_KeyScheduler', 'AES_KeyExpansion_Core', 'AESKeySched_Logic', 'RoundKeySched_Module','KeyExpansion_Controller', 'AESKeyExpandBlock', 'RoundKey_ExpansionCore', 'KeySchedulerUnit','KeyExpansionTop_Module','AES_KeyExp_Core', 'AES128_KeySched_Core', 'KeyGen_Block', 'KeyExpBlock', 'AES_KeySchedulerUnit', 'KeyExpandProcessor',
    'KeySched_Module','AES_KeyExpand_Gen', 'RoundKey_Expander_Mod', 'KeyExpansion_Engine', 'AES_KeyExp_Controller','KeyExpTop_Module', 'KeyExpansionBlock_Core', 'KeyExp_Controller' ,'AES128_KeyExpansion_Unit','AES128KeyExp','AESKeySchedulerCore','AES_KeyExpand_Top', 'AES_KeyExp_Unit', 'KeyExpansion128_Core', 'AES_KeySched_Unit', 'AES_KeySchedulerBlock', 'KeyExpand_128_Block', 'KeyExpand128_Module', 'AESKeySchedulerBlock', 'AES128_KeyScheduler_Mod',
    'AES_KeyExp_Processor', 'KeyExp_Module128','AES128KeySched', 'AES_KeyExpansion_Unit', 'KeySched_Expander', 'AES_KeySched_Block','AES128_KeyExpansionBlock','AES128_KeyExp_Processor', 'AES_KeyExp_Gen_Block', 'KeySched_Top',
    'KeyExp_Gen_Block', 'KeyExpansionRoutine_Mod', 'KeySchedulerProcessor'
]
RSA_LIST_TOP = [
    "rsa_top", "rsa_core","rsa_engine", "rsa_main", "rsa_unit","rsa_top_module",    "rsa_top_level","rsa_wrapper", "rsa_controller", "rsa_processor", "modexp_top" , "modexp_core", "montgomery_core","montgomery_top","rsa_datapath","rsa_pipeline", "rsa_compute", "rsa_main_core", "rsa_block",
    "rsa_modexp", "rsa_montgomery", "exp_core", "exp_unit"
]

SHA_LIST_TOP = [
    "top_main", "sha_top","sha_core", "sha256_top", "sha256_core" , "sha2_core","sha512_top", "sha512_core", "sha_engine","sha_main", "sha_unit","sha_controller" , "sha_wrapper", "sha_top_level", "sha_datapath", "sha_pipeline","sha_round", "sha_sched", "message_schedule", "sha_transform"
]
FSM_LIST_TOP = [
    "fsm", "fsm_module", "state_machine", "ctrl_fsm", "control_fsm" , "controller_fsm","fsm_core", "fsm_top", "fsm_unit", "state_ctrl", "state_ctrl_fsm", "controller", "ctrl", "state", "next_state_logic","fsm_logic"
]


IP_REGISTRY: Dict[str, Dict[str, Dict[str, List[str]]]] = {
    "AES": {"functions": {
        "AddRoundKey": AES_LIST_ADDROUNDKEY,
        "SBox": AES_LIST_SBOX,
        "ShiftRows": AES_LIST_SHIFTROWS,
        "KeyExpansion": AES_LIST_KEYEXPANSION,
    }},
    "RSA": {"functions": {"top_module": RSA_LIST_TOP}},
    "SHA": {"functions": {"top_module": SHA_LIST_TOP}},
    "FSM": {"functions": {"fsm_module": FSM_LIST_TOP}},
}

def sorensen_dice_coefficient(str1: str, str2: str) -> float:
    s1 = str(str1)
    s2 =str(str2)
    b1 =set(s1[i:i + 2] for i in range(max(0, len(s1) - 1)))
    b2 =set(s2[i:i + 2] for i in range(max(0, len(s2) - 1)))
    tot =len(b1) + len(b2)
    if tot ==  0:
        return 0.0
    return 2 * len(b1 & b2) / tot

def _normalize_port_direction(raw: Any) -> str:
    if raw is None:
        return ""
    s= str(raw).strip().lower()
    if s in {"in", "input","portdirection.in","dir_in"}:
        return "input"
    if s in {"out", "output","portdirection.out", "dir_out"}:
        return "output"
    if s in { "inout", "portdirection.inout","ref","dir_inout"}:
        return "inout"
    if isinstance(raw, dict):
        k=str(raw.get("kind", "")).lower()
        if k == "in":
            return "input"
        if k =="out":
            return "output"
        if k == "inout":
            return "inout"
    return ""
def _port_direction_from_member(member: Dict[str, Any]) -> str:
    for key in ("direction","dir","portDirection","io"):
        if key in member:
            d = _normalize_port_direction(member.get(key))
            if d:
                    return d
    decl = member.get("decl", {})
    if isinstance(decl, dict):
        for key in ("direction","dir","portDirection","io"):
            if key in decl:
                d = _normalize_port_direction(decl.get(key))
                if d:
                        return d
    return ""


def _qualified_tail_name(name: Any) -> str:
    s = str(name or "")
    if " " in s:
        s = s.split(" ")[-1]
    if "::" in s:
        s = s.split("::")[-1]
    if "." in s:
        s = s.split(".")[-1]
    return s.strip()


def _parse_sv_int_literal(lit: str) -> Tuple[Optional[int], Optional[int]]:
    s = (lit or "").strip().lower().replace("_", "")
    if not s:
        return None, None

    if re.fullmatch(r"'\s*[01]", s):
        return None, 0 if s.endswith("0") else None

    m = re.fullmatch(r"(?:(\d+)\s*)?'?\s*([s]?)\s*([bodh])\s*([0-9a-fxz]+)", s)
    if not m:
        try:
            v = int(s, 0)
            return None, v
        except Exception:
            return None, None

    w = int(m.group(1)) if m.group(1) else None
    base = m.group(3)
    digits = m.group(4)

    if any(ch in digits for ch in ("x", "z")):
        return w, None

    try:
        if base == "b":
            v = int(digits, 2)
        elif base == "o":
            v = int(digits, 8)
        elif base == "d":
            v = int(digits, 10)
        elif base == "h":
            v = int(digits, 16)
        else:
            v = None
    except Exception:
        v = None

    return w, v


def _enum_width_from_members(members: List[Dict[str, Any]]) -> Optional[int]:
    widths = []
    for mem in members:
        v = mem.get("value")
        if not v:
            continue
        w, _ = _parse_sv_int_literal(str(v))
        if w:
            widths.append(w)
    return max(widths) if widths else None


def _parse_enum_decl_text(s: str) -> Tuple[List[Dict[str, Any]], Optional[int]]:
    txt = (s or "").strip()
    m = re.search(r"\{([\s\S]*?)\}", txt)
    if not m:
        return [], None

    body = m.group(1)

    width_guess = None
    rm = re.search(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", txt[: m.start(0)])
    if rm:
        try:
            a = int(rm.group(1))
            b = int(rm.group(2))
            width_guess = abs(a - b) + 1
        except Exception:
            width_guess = None

    parts = [p.strip() for p in body.split(",") if p.strip()]
    members: List[Dict[str, Any]] = []
    for p in parts:
        p2 = re.split(r"//|/\*", p, maxsplit=1)[0].strip()
        if not p2:
            continue
        if "=" in p2:
            name, val = p2.split("=", 1)
            members.append({"name": _qualified_tail_name(name), "value": val.strip()})
        else:
            members.append({"name": _qualified_tail_name(p2), "value": ""})

    if width_guess is None:
        width_guess = _enum_width_from_members(members)

    return members, width_guess


def _extract_enum_typedef_ref(type_str: str) -> Optional[str]:
    s = str(type_str or "").strip()
    if not s:
        return None

    if "}" in s:
        suffix = s.split("}")[-1].strip()
        tail = _qualified_tail_name(suffix)
        return tail or None

    tail = _qualified_tail_name(s)
    return tail or None


def _infer_type_width_enumref(type_info: Any,
                             enum_types: Dict[str, Dict[str, Any]]) -> Tuple[str, int, Optional[str]]:
    out_type = "logic"
    out_w = 1
    enum_ref: Optional[str] = None

    if isinstance(type_info, dict):
        packed = type_info.get("packedRange")
        if isinstance(packed, dict) and "left" in packed and "right" in packed:
            try:
                out_w = abs(int(packed["left"]) - int(packed["right"])) + 1
            except Exception:
                out_w = 1

        kind = str(type_info.get("kind", "")).lower()
        base = str(type_info.get("base", "") or "").lower()

        cand = None
        for k in ("name", "type", "ref", "target"):
            v = type_info.get(k)
            if isinstance(v, str) and v:
                cand = v
                break
        if cand:
            er = _extract_enum_typedef_ref(cand)
            if er and er in enum_types:
                enum_ref = er

        if "enum" in kind or "enum" in base or enum_ref:
            out_type = "enum"
            if enum_ref and isinstance(enum_types.get(enum_ref, {}).get("width"), int):
                out_w = int(enum_types[enum_ref]["width"])
            else:
                members = []
                for k in ("members", "items", "enumerators"):
                    v = type_info.get(k)
                    if isinstance(v, list) and v:
                        for it in v:
                            if isinstance(it, dict) and "name" in it:
                                members.append({"name": _qualified_tail_name(it.get("name")),
                                                "value": str(it.get("value", "")).strip()})
                            else:
                                members.append({"name": _qualified_tail_name(it), "value": ""})
                w2 = _enum_width_from_members(members) if members else None
                if w2:
                    out_w = int(w2)

        if out_type == "logic":
            if base:
                out_type = base
            elif kind:
                out_type = kind

        return out_type, out_w, enum_ref

    s = str(type_info or "").strip()
    if not s:
        return out_type, out_w, None

    tail = _extract_enum_typedef_ref(s)
    if tail and tail in enum_types:
        enum_ref = tail
        out_type = "enum"
        w2 = enum_types[tail].get("width")
        if isinstance(w2, int) and w2 > 0:
            out_w = int(w2)
        return out_type, out_w, enum_ref

    sl = s.lower()

    if "enum" in sl and "{" in sl and "}" in sl:
        mems, w = _parse_enum_decl_text(s)
        out_type = "enum"
        if isinstance(w, int) and w > 0:
            out_w = int(w)
        else:
            out_w = int(_enum_width_from_members(mems) or 1)

        enum_ref = _extract_enum_typedef_ref(s)
        if enum_ref and enum_ref in enum_types:
            w2 = enum_types[enum_ref].get("width")
            if isinstance(w2, int) and w2 > 0:
                out_w = int(w2)
        else:
            enum_ref = None

        return out_type, out_w, enum_ref

    rm = re.search(r"\[\s*(\d+)\s*:\s*(\d+)\s*\]", s)
    if rm:
        try:
            a = int(rm.group(1))
            b = int(rm.group(2))
            out_w = abs(a - b) + 1
        except Exception:
            out_w = 1

    bm = re.match(r"^\s*([a-zA-Z_]\w*)", s)
    if bm:
        out_type = bm.group(1).lower()

    return out_type, out_w, None


def _infer_type_and_width(type_info: Any, enum_types: Dict[str, Dict[str, Any]]) -> Tuple[str, int]:
    t, w, _ = _infer_type_width_enumref(type_info, enum_types)
    return t, w


class PDGBuilder:
    def __init__(self, module_node=None):
        self.pdg: Dict[str, Dict[str, Any]] = {}
        self.input_list: List[str] = []
        self.visited = {}
        self.max_depth_cache = {}
        self.modules_dict: Dict[str, Any] = {}
        self.variables: List[Dict[str, Any]] = []
        self.current_module_name = ""
        self.module_instances = {}

        self.enum_types: Dict[str, Dict[str, Any]] = {}
        self.enum_member_info: Dict[str, Dict[str, Any]] = {}

        if module_node:
            self.build_pdg(module_node)

    def find_paths_to_inputs(self, start_node):
        paths = []
        stack = [(start_node, [start_node])]
        while stack:
            (node, path) = stack.pop()
            if node in self.input_list:
                paths.append(path)
            else:
                for neighbor in self.pdg[node]["connections"]:
                    if neighbor not in path:
                        stack.append((neighbor, path + [neighbor]))
        return paths

    def calculate_centroid(self):
        centroid_scores = {}
        for node_name, node_data in self.pdg.items():
            if node_data["type"] == "Variable":
                paths = self.find_paths_to_inputs(node_name)
                total_weight = 0
                for path in paths:
                    path_weight = 0
                    for n in path:
                        operators = self.pdg[n].get("operators", {})
                        path_weight += sum(operators.values())
                    total_weight += path_weight
                centroid_scores[node_name] = total_weight

        max_score = max(centroid_scores.values()) if centroid_scores else 0
        if max_score == 0:
            for node_name in centroid_scores:
                self.pdg[node_name]["centroid"] = 0
        else:
            for node_name, score in centroid_scores.items():
                self.pdg[node_name]["centroid"] = score / max_score

    def _append_variable(self, var_map: Dict[str, Dict[str, Any]], record: Dict[str, Any], is_port: bool):
        name = record["name"]
        if not name:
            return

        if name not in var_map:
            var_map[name] = record
            return

        existing = var_map[name]
        existing_is_port = existing.get("_kind") == "Port" and existing.get("type") in ("input", "output", "inout")

        if is_port:
            var_map[name] = record
            return

        if existing_is_port:
            try:
                ew = int(existing.get("length", 1))
                nw = int(record.get("length", 1))
                if nw > ew:
                    existing["length"] = nw
            except Exception:
                pass
            if record.get("enum_type") and not existing.get("enum_type"):
                existing["enum_type"] = record.get("enum_type")
            return

        ex_t = str(existing.get("type", "")).lower()
        nw_t = str(record.get("type", "")).lower()

        if ex_t in ("unknown", "", "logic") and nw_t not in ("unknown", "", "logic"):
            var_map[name] = record
            return

        try:
            ew = int(existing.get("length", 1))
        except Exception:
            ew = 1
        try:
            nw = int(record.get("length", 1))
        except Exception:
            nw = 1

        if nw > ew:
            existing["length"] = nw

        if nw_t == "enum" and ex_t != "enum":
            existing["type"] = "enum"

        if record.get("enum_type") and not existing.get("enum_type"):
            existing["enum_type"] = record.get("enum_type")

    def _collect_enum_typedefs(self, members: List[Dict[str, Any]]) -> None:
        for member in members or []:
            kind = str(member.get("kind", "") or "")
            if kind not in ("TypeAlias", "Typedef", "TypeDef", "TypeDefDeclaration"):
                continue

            tname = _qualified_tail_name(member.get("name", ""))

            target = member.get("target")
            txt = ""
            if isinstance(target, str):
                txt = target
            elif isinstance(target, dict):
                for k in ("text", "decl", "name"):
                    if k in target and isinstance(target.get(k), str):
                        txt = target.get(k)
                        break
                if not txt:
                    txt = json.dumps(target)
            else:
                txt = str(member.get("type", "") or "")

            if "enum" not in str(txt).lower():
                continue

            mems, w = _parse_enum_decl_text(txt)
            if not mems:
                continue

            enum_name = tname or f"anon_enum_{id(member)}"
            width = int(w) if isinstance(w, int) and w > 0 else int(_enum_width_from_members(mems) or 1)

            self.enum_types[enum_name] = {"width": width, "members": mems}

            for m in mems:
                mn = _qualified_tail_name(m.get("name", ""))
                if not mn:
                    continue
                prev = self.enum_member_info.get(mn, {})
                prev_w = int(prev.get("length", 0) or 0)
                if width >= prev_w:
                    self.enum_member_info[mn] = {"name": mn, "type": "enum_member", "length": width, "enum_type": enum_name}

    def extract_variables_and_statements_from_module(self, node):
        variables_map: Dict[str, Dict[str, Any]] = {}
        statements = []
        instances = []
        if not node:
            return [], statements, instances

        members = node.get("members", []) or []

        self._collect_enum_typedefs(members)

        for em_name, rec in self.enum_member_info.items():
            rec2 = {
                "name": em_name,
                "type": rec.get("type", "enum_member"),
                "length": rec.get("length", 1),
                "enum_type": rec.get("enum_type", ""),
                "_kind": "EnumMember",
            }
            self._append_variable(variables_map, rec2, is_port=False)

        for member in members:
            kind = member.get("kind", "")

            if kind in ["Port", "Variable", "Net", "Parameter"]:
                var_name = _qualified_tail_name(member.get("name"))
                if not var_name:
                    continue

                port_dir = _port_direction_from_member(member) if kind == "Port" else ""
                raw_type = member.get("type", "")

                inferred_type, inferred_w, enum_ref = _infer_type_width_enumref(raw_type, self.enum_types)

                if kind == "Net" and not port_dir:
                    net_type = ""
                    nt = member.get("netType")
                    if isinstance(nt, dict):
                        net_type = str(nt.get("name", "") or "").strip().lower()
                    if net_type:
                        inferred_type = net_type
                    elif inferred_type in ("logic", "unknown", ""):
                        inferred_type = "wire"

                final_type = port_dir or inferred_type
                length = inferred_w if isinstance(inferred_w, int) and inferred_w > 0 else 1

                rec = {
                    "name": var_name,
                    "type": final_type,
                    "length": length,
                    "enum_type": enum_ref or "",
                    "_kind": kind,
                }
                self._append_variable(variables_map, rec, is_port=(kind == "Port"))

                if "initializer" in member:
                    assignment = {
                        "kind": "VariableInitialization",
                        "left": {"kind": "NamedValue", "symbol": f"{member.get('addr', '')} {var_name}"},
                        "right": member["initializer"],
                    }
                    statements.append(assignment)

            elif kind in ["ContinuousAssign"]:
                statements.append(member)

            elif kind in ["ProceduralBlock", "Always", "Initial"]:
                statements.append(member)
                self.extract_variables_from_statement(member, list(variables_map.values()))

            elif kind == "Instance":
                instances.append(member)

            elif kind == "TransparentMember":
                nm = _qualified_tail_name(member.get("name", ""))
                if nm:
                    lit = member.get("value") or member.get("initializer") or ""
                    w, _ = _parse_sv_int_literal(str(lit))
                    width = int(w) if w else 1
                    rec = {"name": nm, "type": "enum_member", "length": width, "_kind": "TransparentMember"}
                    self._append_variable(variables_map, rec, is_port=False)

        variables = []
        for v in variables_map.values():
            v.pop("_kind", None)
            variables.append(v)

        return variables, statements, instances

    def extract_variables_from_statement(self, stmt, variables):
        if not isinstance(stmt, dict):
            return
        kind = stmt.get("kind", "")
        if kind in ["ProceduralBlock", "Block", "Timed", "List", "StatementBlock", "Conditional", "ForLoop"]:
            body = stmt.get("body") or stmt.get("list") or stmt.get("stmts") or []
            if isinstance(body, dict):
                self.extract_variables_from_statement(body, variables)
            elif isinstance(body, list):
                for s in body:
                    self.extract_variables_from_statement(s, variables)

        elif kind == "VariableDeclaration":
            var_name = _qualified_tail_name(stmt.get("symbol", ""))
            if var_name and not any(var["name"] == var_name for var in variables):
                t, w, enum_ref = _infer_type_width_enumref(stmt.get("type", "Unknown"), self.enum_types)
                variables.append({"name": var_name, "type": t or "Unknown", "length": w or 1, "enum_type": enum_ref or ""})

        elif kind == "ExpressionStatement":
            expr = stmt.get("expr", {})
            if expr.get("kind") == "Assignment":
                left = expr.get("left", {})
                var_name = self.extract_target_name(left)
                if var_name and not any(var["name"] == var_name for var in variables):
                    variables.append({"name": var_name, "type": "Unknown", "length": 1})

        elif kind == "Conditional":
            ifTrue = stmt.get("ifTrue", {})
            ifFalse = stmt.get("ifFalse", {})
            self.extract_variables_from_statement(ifTrue, variables)
            self.extract_variables_from_statement(ifFalse, variables)

        elif kind == "Case":
            items = stmt.get("items", [])
            for item in items:
                self.extract_variables_from_statement(item.get("stmt", {}), variables)
            default_case = stmt.get("defaultCase")
            if default_case:
                self.extract_variables_from_statement(default_case, variables)

    def extract_sources(self, expression):
        sources = {"names": [], "ops": []}

        if isinstance(expression, list):
            for expr in expression:
                nested_sources = self.extract_sources(expr)
                sources["names"].extend(nested_sources["names"])
                sources["ops"].extend(nested_sources["ops"])
            return sources

        if not isinstance(expression, dict):
            return sources

        kind = expression.get("kind", "")
        if kind == "NamedValue":
            if "symbol" in expression:
                sources["names"].append(_qualified_tail_name(expression["symbol"]))

        elif kind == "BinaryOp":
            operator = expression.get("op") or expression.get("operator")
            if operator:
                sources["ops"].append(operator)
            left_sources = self.extract_sources(expression.get("left", {}))
            right_sources = self.extract_sources(expression.get("right", {}))
            sources["names"].extend(left_sources["names"] + right_sources["names"])
            sources["ops"].extend(left_sources["ops"] + right_sources["ops"])

        elif kind == "UnaryOp":
            operator = expression.get("op") or expression.get("operator")
            if operator:
                sources["ops"].append(operator)
            operand_sources = self.extract_sources(expression.get("operand", {}))
            sources["names"].extend(operand_sources["names"])
            sources["ops"].extend(operand_sources["ops"])

        elif kind == "ConditionalOp":
            cond_sources = self.extract_sources(expression.get("cond", {}))
            true_sources = self.extract_sources(expression.get("left", {}))
            false_sources = self.extract_sources(expression.get("right", {}))
            sources["names"].extend(cond_sources["names"] + true_sources["names"] + false_sources["names"])
            sources["ops"].extend(cond_sources["ops"] + true_sources["ops"] + false_sources["ops"])

        elif kind == "ElementSelect":
            value_sources = self.extract_sources(expression.get("value", {}))
            selector_sources = self.extract_sources(expression.get("selector", {}))
            sources["names"].extend(value_sources["names"] + selector_sources["names"])

        elif kind == "Conversion":
            operand_sources = self.extract_sources(expression.get("operand", {}))
            sources["names"].extend(operand_sources["names"])
            sources["ops"].extend(operand_sources["ops"])

        return sources

    def extract_condition_str(self, expr):
        kind = expr.get("kind", "")
        if kind == "NamedValue":
            return _qualified_tail_name(expr.get("symbol", ""))
        elif kind == "BinaryOp":
            left = self.extract_condition_str(expr["left"])
            right = self.extract_condition_str(expr["right"])
            op = expr.get("op", "")
            return f"({left} {op} {right})"
        elif kind == "UnaryOp":
            operand = self.extract_condition_str(expr["operand"])
            op = expr.get("op", "")
            return f"({op}{operand})"
        elif kind == "IntegerLiteral":
            return str(expr.get("value", ""))
        else:
            return ""

    def extract_target_name(self, expr):
        kind = expr.get("kind", "")
        if kind in ["NamedValue", "Variable"]:
            if "symbol" in expr:
                return _qualified_tail_name(expr["symbol"])
        elif kind in ["ElementSelect", "RangeSelect"]:
            if "value" in expr:
                return self.extract_target_name(expr["value"])
        elif kind == "MemberAccess":
            if "parent" in expr and "member" in expr:
                base_name = self.extract_target_name(expr["parent"])
                member_name = expr["member"]
                return f"{base_name}.{member_name}"
        return None

    def ensure_nodes_in_pdg(self, nodes):
        for node in nodes:
            if node not in self.pdg:
                self.pdg[node] = {
                    "name": node,
                    "type": "Variable",
                    "connections": [],
                    "conditions": [],
                    "operators": Counter(),
                    "details": {},
                }

    def update_pdg(self, target, sources, conditions, operators=None):
        if target is None:
            return
        if target not in self.pdg:
            self.pdg[target] = {
                "name": target,
                "type": "Variable",
                "connections": [],
                "conditions": [],
                "operators": Counter(),
                "details": {},
            }
        self.ensure_nodes_in_pdg(sources)
        self.pdg[target]["connections"].extend(sources)
        self.pdg[target]["connections"] = sorted(set(self.pdg[target]["connections"]))
        self.pdg[target]["conditions"].extend(conditions)
        self.pdg[target]["conditions"] = sorted(set(self.pdg[target]["conditions"]))
        if operators is not None:
            self.pdg[target]["operators"].update(operators)

    def add_connections_conditional(self, node, current_conditions, parent_nodes):
        if node.get("kind") != "Conditional":
            self.process_statement(node, current_conditions, parent_nodes)
            return

        conditions = node.get("conditions", [])
        for condition in conditions:
            condition_expr = condition.get("expr", {})
            condition_str = self.extract_condition_str(condition_expr)

            cond_sources = self.extract_sources(condition_expr)
            cond_node_name = f"cond_{id(condition)}"
            self.ensure_nodes_in_pdg(cond_sources["names"])
            self.ensure_nodes_in_pdg(parent_nodes)

            self.pdg[cond_node_name] = {
                "name": cond_node_name,
                "type": "Condition",
                "connections": list(set(cond_sources["names"] + parent_nodes)),
                "conditions": [],
                "operators": Counter(),
                "details": {},
            }

            ifTrue = node.get("ifTrue", {})
            next_parents = parent_nodes + [cond_node_name]
            if ifTrue:
                self.process_statement(ifTrue, current_conditions + [condition_str], next_parents)

            ifFalse = node.get("ifFalse", {})
            if ifFalse:
                condition_negation_str = f"!({condition_str})"
                self.process_statement(ifFalse, current_conditions + [condition_negation_str], next_parents)

    def add_connections_case(self, case_expr, current_conditions, parent_nodes):
        case_cond = case_expr.get("expr", {})
        case_expr_str = self.extract_condition_str(case_cond)
        case_sources = self.extract_sources(case_cond)
        case_node_name = f"case_{id(case_expr)}"
        self.ensure_nodes_in_pdg(case_sources["names"])
        self.ensure_nodes_in_pdg(parent_nodes)

        self.pdg[case_node_name] = {
            "name": case_node_name,
            "type": "Condition",
            "connections": case_sources["names"] + parent_nodes,
            "conditions": [],
            "operators": Counter(),
            "details": {},
        }

        items = case_expr.get("items", [])
        for item in items:
            item_exprs = item.get("expressions", [])
            for expr in item_exprs:
                expr_str = self.extract_condition_str(expr)
                condition_str = f"({case_expr_str} == {expr_str})"
                self.process_statement(item.get("stmt", {}), current_conditions + [condition_str],
                                       parent_nodes + [case_node_name])

        default_case = case_expr.get("defaultCase")
        if default_case:
            condition_str = f"default({case_expr_str})"
            self.process_statement(default_case, current_conditions + [condition_str],
                                   parent_nodes + [case_node_name])

    def process_statement(self, node, current_conditions, parent_nodes=None):
        if parent_nodes is None:
            parent_nodes = []
        kind = node.get("kind", "")

        if kind == "ExpressionStatement":
            expr = node.get("expr", {})
            if "left" in expr and "right" in expr:
                target = self.extract_target_name(expr["left"])
                sources = self.extract_sources(expr["right"])
                operators = Counter(sources["ops"])
                connections = sources["names"] + parent_nodes
                self.update_pdg(target, connections, current_conditions, operators)

                if "symbol" in expr.get("left", {}):
                    target2 = _qualified_tail_name(expr["left"]["symbol"])
                    self.update_pdg(target2, connections, current_conditions, operators)

        elif kind == "ContinuousAssign":
            assignment = node.get("assignment", {})
            if "left" in assignment and "right" in assignment:
                target = self.extract_target_name(assignment["left"])
                sources = self.extract_sources(assignment["right"])
                operators = Counter(sources["ops"])
                connections = sources["names"] + parent_nodes
                self.update_pdg(target, connections, current_conditions, operators)

        elif kind == "Block":
            body = node.get("body")
            if isinstance(body, dict):
                self.process_statement(body, current_conditions, parent_nodes)
            elif isinstance(body, list):
                for stmt in body:
                    self.process_statement(stmt, current_conditions, parent_nodes)

        elif kind == "List":
            for stmt in node.get("list", []):
                self.process_statement(stmt, current_conditions, parent_nodes)

        elif kind == "Conditional":
            self.add_connections_conditional(node, current_conditions, parent_nodes)

        elif kind == "Case":
            self.add_connections_case(node, current_conditions, parent_nodes)

        elif kind == "Timed":
            if "stmt" in node:
                self.process_statement(node["stmt"], current_conditions, parent_nodes)

        elif kind == "ProceduralBlock":
            self.process_statement(node.get("body", {}), current_conditions, parent_nodes)

        elif kind == "Assignment":
            left = node.get("left", {})
            target = _qualified_tail_name(left.get("symbol", "")) if "symbol" in left else None
            sources = self.extract_sources(node.get("right", {}))
            connections = sources["names"] + parent_nodes
            self.update_pdg(target, connections, current_conditions)

        elif kind == "ForLoop":
            initializer = node.get("init")
            condition = node.get("cond")
            increment = node.get("inc")
            body = node.get("body")

            if initializer:
                self.process_statement(initializer, current_conditions, parent_nodes)
            if condition:
                _ = self.extract_sources(condition)
            if increment:
                _ = self.extract_sources(increment)
            if body:
                if isinstance(body, list):
                    for stmt in body:
                        self.process_statement(stmt, current_conditions, parent_nodes)
                elif isinstance(body, dict):
                    self.process_statement(body, current_conditions, parent_nodes)

    def handle_node_types(self, node):
        kind = node.get("kind", "")
        if kind in ["Port", "Variable", "Net", "Parameter"]:
            name = _qualified_tail_name(node.get("name", ""))
            if kind == "Port":
                direction = _port_direction_from_member(node)
                if direction in ("input", "inout"):
                    self.input_list.append(name)
            if name and (name not in self.pdg):
                self.pdg[name] = {
                    "name": name,
                    "type": "Variable",
                    "connections": [],
                    "conditions": [],
                    "operators": Counter(),
                    "details": node,
                }

        elif kind == "ContinuousAssign":
            assignment = node.get("assignment", {})
            if "left" in assignment and "right" in assignment:
                target = self.extract_target_name(assignment["left"])
                sources = self.extract_sources(assignment["right"])
                self.ensure_nodes_in_pdg(sources["names"])
                if target:
                    self.update_pdg(target, sources["names"], [])

        if "body" in node and kind not in ["ProceduralBlock", "Block", "Conditional", "List", "Timed"]:
            if isinstance(node["body"], dict):
                self.handle_node_types(node["body"])
            elif isinstance(node["body"], list):
                for item in node["body"]:
                    self.handle_node_types(item)
        if "members" in node:
            for member in node["members"]:
                self.handle_node_types(member)
        if "stmts" in node:
            for stmt in node["stmts"]:
                self.handle_node_types(stmt)
        if "expr" in node and isinstance(node.get("expr"), dict) and node["expr"].get("kind", "") == "ConditionalOp":
            self.process_statement(node, [])

    def calculate_dependence_depth(self, start_name):
        def dfs(name):
            if name in self.max_depth_cache:
                return self.max_depth_cache[name]
            if name in self.visited:
                return 0
            self.visited.add(name)
            max_depth = 0
            if name in self.pdg:
                node = self.pdg[name]
                connections = node["connections"]
                conditions = node["conditions"]
                count_cnd = sum(1 for item in connections if str(item).startswith("cond_"))
                count_cse = sum(1 for item in connections if str(item).startswith("case_"))

                if len(conditions) >= 1:
                    for cond in conditions:
                        cond_depth = dfs(cond)
                        if cond_depth > max_depth:
                            max_depth = cond_depth
                    if len(connections) >= 1:
                        for conn in connections:
                            if str(conn).startswith("cond_"):
                                conn_depth = 0
                            else:
                                conn_depth = dfs(conn)
                            if conn_depth > max_depth:
                                max_depth = conn_depth
                else:
                    for conn in connections:
                        if str(conn).startswith("cond_"):
                            max_depth = max(max_depth, 0)
                        else:
                            conn_depth = dfs(conn)
                            if conn_depth > max_depth:
                                max_depth = conn_depth

                max_depth = count_cnd + max_depth + count_cse

            self.visited.remove(name)
            depth = max_depth + 1
            self.max_depth_cache[name] = depth
            return depth

        self.visited = set()
        self.max_depth_cache = {}
        total_depth = dfs(start_name)
        return total_depth - 1

    def build_pdg(self, module_node):
        variables, statements, instances = self.extract_variables_and_statements_from_module(module_node)
        self.variables = variables

        for var in variables:
            name = var["name"]
            var_type = str(var.get("type", "")).lower()
            if var_type in ("input", "inout"):
                self.input_list.append(name)
            if name not in self.pdg:
                self.pdg[name] = {
                    "name": name,
                    "type": "Variable",
                    "connections": [],
                    "conditions": [],
                    "operators": Counter(),
                    "details": var,
                }

        for stmt in statements:
            self.process_statement(stmt, [])

        for instance in instances:
            self.handle_node_types(instance)

        self.calculate_centroid()

    def aggregate_operators(self):
        for node_name in list(self.pdg.keys()):
            self._visited_agg = set()
            operators = self.get_aggregated_operators(node_name)
            self.pdg[node_name]["aggregated_operators"] = operators

    def get_aggregated_operators(self, node_name):
        if "aggregated_operators" in self.pdg[node_name]:
            return self.pdg[node_name]["aggregated_operators"]
        if not hasattr(self, "_visited_agg"):
            self._visited_agg = set()
        if node_name in self._visited_agg:
            return Counter()
        self._visited_agg.add(node_name)

        operators = Counter(self.pdg[node_name]["operators"])
        for conn in self.pdg[node_name]["connections"]:
            if conn in self.pdg:
                conn_operators = self.get_aggregated_operators(conn)
                operators.update(conn_operators)
        self.pdg[node_name]["aggregated_operators"] = operators
        return operators

    def main(self, json_data, module_name):
        def build_modules_dict_local(node, out: Dict[str, Any]):
            if isinstance(node, dict):
                if node.get("kind") == "InstanceBody":
                    mname = node.get("name")
                    if mname:
                        out[mname] = node
                    for member in node.get("members", []):
                        build_modules_dict_local(member, out)
                else:
                    for _, value in node.items():
                        if isinstance(value, (dict, list)):
                            build_modules_dict_local(value, out)
            elif isinstance(node, list):
                for item in node:
                    build_modules_dict_local(item, out)

        self.modules_dict = {}
        build_modules_dict_local(json_data, self.modules_dict)
        module_node = self.modules_dict.get(module_name)
        if module_node is None:
            print(f"module '{module_name}' not found in the JSON data.")
            return

        self.current_module_name = module_name
        self.build_pdg(module_node)
        self.aggregate_operators()

        for name, node in list(self.pdg.items()):
            if node["type"] in ("Variable", "Condition"):
                self.visited = {}
                self.max_depth_cache = {}
                depth = self.calculate_dependence_depth(name)
                self.pdg[name]["depth"] = depth

    def get_pdg_depths(self):
        pdg_depths = {}
        for name, node in self.pdg.items():
            if node["type"] in ("Variable", "Condition"):
                pdg_depths[name] = node.get("depth", "N/A")
        return pdg_depths

    def get_variables(self):
        return self.variables


def build_modules_dict(node: Any, modules_dict: Dict[str, Any]) -> None:
    if isinstance(node, dict):
        if node.get("kind") == "InstanceBody":
            module_name = node.get("name")
            if module_name:
                modules_dict[module_name] = node
            for member in node.get("members", []):
                build_modules_dict(member, modules_dict)
        else:
            for _, value in node.items():
                if isinstance(value, (dict, list)):
                    build_modules_dict(value, modules_dict)
    elif isinstance(node, list):
        for item in node:
            build_modules_dict(item, modules_dict)


def _safe_int(x) -> Optional[int]:
    try:
        if x is None:
            return None
        if isinstance(x, bool):
            return int(x)
        if isinstance(x, (int, float)):
            if str(x) == "nan":
                return None
            return int(x)
        s = str(x).strip()
        if not s or s.lower() == "nan":
            return None
        return int(float(s))
    except Exception:
        return None


def export_pdg_pack_json(*,
                         out_path: str,
                         family: str,
                         design_file: str,
                         function: str,
                         module: str,
                         builder: PDGBuilder,
                         variable_info: Dict[str, Dict[str, Any]]) -> None:
    nodes = {}
    reverse_edges = defaultdict(set)

    for name, node in builder.pdg.items():
        conns = [str(c) for c in node.get("connections", [])]
        for c in conns:
            reverse_edges[c].add(name)

        vmeta = variable_info.get(name, {})
        if not vmeta and node.get("type") == "Condition":
            vmeta = {"type": "Condition", "length": None}

        enum_type = ""
        if isinstance(vmeta, dict) and str(vmeta.get("type", "")).lower() == "enum":
            enum_type = str(vmeta.get("enum_type", "") or "")

        nodes[name] = {
            "node_type": node.get("type", ""),
            "connections": conns,
            "depth": _safe_int(node.get("depth")),
            "centroid": float(node.get("centroid", 0.0) or 0.0),
            "operators": dict(node.get("operators", {})),
            "aggregated_operators": dict(node.get("aggregated_operators", {})),
            "var_type": str(vmeta.get("type", "")) if isinstance(vmeta, dict) else "",
            "bit_width": _safe_int(vmeta.get("length")) if isinstance(vmeta, dict) else None,
            "enum_type": enum_type,
        }

    inputs = sorted(set([str(x) for x in builder.input_list]))
    outputs = sorted(set([n for n, v in variable_info.items()
                          if isinstance(v, dict) and str(v.get("type", "")).lower() == "output"]))

    enums_out = {}
    for ename, edef in (builder.enum_types or {}).items():
        enums_out[str(ename)] = {
            "width": int(edef.get("width", 1) or 1),
            "members": list(edef.get("members", []) or []),
        }

    pack = {
        "family": family,
        "design_file": design_file,
        "function": function,
        "module": module,
        "inputs": inputs,
        "outputs": outputs,
        "enums": enums_out,
        "nodes": nodes,
        "reverse_edges": {k: sorted(list(v)) for k, v in reverse_edges.items()},
    }

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(pack, f, indent=2)


def main():
    parser = argparse.ArgumentParser(description="PDG batch analyzer (AES/RSA/SHA/FSM)")
    parser.add_argument("--family", required=True, choices=list(IP_REGISTRY.keys()),
                        help="IP family to use (AES, RSA, SHA, FSM)")
    parser.add_argument("--design_folder", required=True,
                        help="Folder containing *.json AST files")

    parser.add_argument("--out_dir", default=".",
                        help="Where to write per-design Excel metric files (default: current dir)")
    parser.add_argument("--pdg_json_dir", default=None,
                        help="Optional: where to write PDG pack JSON files (one per design+function)")

    parser.add_argument("--summary", default=None,
                        help="Optional output filename for the summary workbook (default: Summary_<family>.xlsx)")

    args = parser.parse_args()

    family = args.family
    functions = IP_REGISTRY[family]["functions"]
    design_folder = args.design_folder
    out_dir = args.out_dir
    pdg_json_dir = args.pdg_json_dir

    if not os.path.isdir(design_folder):
        raise FileNotFoundError(f"Design folder not found: {design_folder}")
    design_files = [f for f in os.listdir(design_folder) if f.endswith(".json")]
    if not design_files:
        print(f"no *.json files found in {design_folder}. Nothing to do.")
        return

    os.makedirs(out_dir, exist_ok=True)
    if pdg_json_dir:
        os.makedirs(pdg_json_dir, exist_ok=True)

    summary_data: Dict[str, Dict[str, str]] = {}

    for design_file in design_files:
        json_file_path = os.path.join(design_folder, design_file)
        try:
            with open(json_file_path, "r", encoding="utf-8") as f:
                json_data = json.load(f)
        except Exception as e:
            print(f"failed to read '{json_file_path}': {e}")
            continue

        modules_dict: Dict[str, Any] = {}
        build_modules_dict(json_data, modules_dict)
        module_names = list(modules_dict.keys())

        design_stem = os.path.splitext(design_file)[0]
        excel_file = f"{design_stem}.xlsx"
        excel_path = os.path.join(out_dir, excel_file)

        workbook = openpyxl.Workbook()
        created_any_sheet = False
        summary_data[design_file] = {}

        for func_name, alias_list in functions.items():
            print(f"\n[INFO]::: Processing list: {func_name} | file: {design_file}")
            similarity_counter = Counter()

            for alias in alias_list:
                highest_similarity = 0.0
                most_similar_name = None
                for module_name in module_names:
                    similarity = sorensen_dice_coefficient(alias.lower(), module_name.lower())
                    if similarity > highest_similarity:
                        highest_similarity = similarity
                        most_similar_name = module_name
                if most_similar_name:
                    similarity_counter[most_similar_name] += 1

            if similarity_counter:
                best_module, count = similarity_counter.most_common(1)[0]
                print(f"most repeated similar module in {func_name}: '{best_module}' (x{count})")
                summary_data[design_file][func_name] = best_module

                module_node = modules_dict.get(best_module)
                if module_node is None:
                    print(f"module '{best_module}' not found in design. Skipping sheet '{func_name}'.")
                    continue

                builder = PDGBuilder(module_node)
                builder.main(json_data, best_module)

                pdg_depths = builder.get_pdg_depths()
                variables = builder.get_variables()
                variable_info = {v["name"]: v for v in variables}

                sheet_title = func_name[:31]
                if sheet_title in workbook.sheetnames:
                    sheet = workbook[sheet_title]
                else:
                    sheet = workbook.create_sheet(title=sheet_title)

                if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
                    try:
                        workbook.remove(workbook["Sheet"])
                    except Exception:
                        pass

                sheet.append(["Variable Name", "Type", "Bit Width", "PDG_Depth", "Num_Operators", "Centroid"])

                for var_name, depth in pdg_depths.items():
                    info = variable_info.get(var_name, None)
                    if info is None:
                        nd = builder.pdg.get(var_name, {})
                        inferred_t = "Condition" if nd.get("type") == "Condition" else "Unknown"
                        info = {"type": inferred_t, "length": "Unknown"}

                    pdg_node = builder.pdg.get(var_name, {})
                    num_ops = sum(pdg_node.get("aggregated_operators", {}).values()) / 2 if pdg_node else 0
                    centroid = pdg_node.get("centroid", 0) if pdg_node else 0

                    tcell = info.get("type", "Unknown")
                    if str(tcell).lower() == "enum" and info.get("enum_type"):
                        tcell = f"enum({info.get('enum_type')})"

                    sheet.append([
                        var_name,
                        tcell,
                        info.get("length", "Unknown"),
                        depth,
                        num_ops,
                        centroid,
                    ])

                created_any_sheet = True

                if builder.enum_types:
                    enum_sheet_title = (func_name + "_enums")[:31]
                    if enum_sheet_title in workbook.sheetnames:
                        es = workbook[enum_sheet_title]
                    else:
                        es = workbook.create_sheet(title=enum_sheet_title)

                    if es.max_row == 1 and es.max_column == 1 and es["A1"].value is None:
                        es.append(["Enum Type", "Width", "Member Name", "Member Value"])

                    for ename, edef in builder.enum_types.items():
                        w = int(edef.get("width", 1) or 1)
                        for mem in (edef.get("members", []) or []):
                            es.append([ename, w, mem.get("name", ""), mem.get("value", "")])

                if pdg_json_dir:
                    out_json = os.path.join(pdg_json_dir, f"pdg__{design_stem}__{func_name}.json")
                    try:
                        export_pdg_pack_json(
                            out_path=out_json,
                            family=family,
                            design_file=design_file,
                            function=func_name,
                            module=best_module,
                            builder=builder,
                            variable_info=variable_info,
                        )
                        print(f"PDG pack saved: {out_json}")
                    except Exception as e:
                        print(f"failed to export PDG pack for {design_stem}:{func_name}: {e}")

            else:
                print(f"no similar modules found for {func_name} in {design_file}.")
                summary_data[design_file][func_name] = "None"

        if not created_any_sheet:
            sheet = workbook.active
            sheet.title = "Info"
            sheet.append(["No function sheets were created for this design."])

        try:
            workbook.save(excel_path)
            print(f"wrote: {excel_path}")
        except Exception as e:
            print(f"failed to save '{excel_path}': {e}")

    summary_excel_file = args.summary or f"Summary_{family}.xlsx"
    summary_path = os.path.join(out_dir, summary_excel_file) if args.summary is None else args.summary

    summary_wb = openpyxl.Workbook()
    summary_sheet = summary_wb.active
    summary_sheet.title = "Summary"
    headers = ["Design File"] + list(functions.keys())
    summary_sheet.append(headers)

    for design_file, funcs in summary_data.items():
        row = [design_file]
        for func_name in functions.keys():
            row.append(funcs.get(func_name, "None"))
        summary_sheet.append(row)

    try:
        summary_wb.save(summary_path)
        print(f"\nsummary saved to::::::: {summary_path}")
    except Exception as e:
        print(f"failed to save summary '{summary_path}': {e}")


if __name__ == "__main__":
    
    main()

