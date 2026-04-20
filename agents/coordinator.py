##coordinator agent
from __future__ import annotations

import os
import re

import logging
import shutil


import subprocess
import importlib
from dataclasses import dataclass
from textwrap import dedent
from typing import Optional, Tuple, Dict

from .sva_generator import generate_sva

from .repair_agent import repair_wrapper

from .log_analyzer import analyze_jg_log, JGResult

from tools.remote_runner import run_verification
from tools.property_loader import load_properties
from .design_analyzer import analyze_design
#from . import static_bridge  



logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)
GENERATED_DIR = os.path.join(".", "generated")
LOGS_DIR = os.path.join(".", "logs")
os.makedirs(GENERATED_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)


@dataclass
class PipelineConfig:
    max_repairs: int= 5


def _load_text(p: str) -> str:
    try:
        with open(p, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return ""
def _write_text(p: str, s: str) -> None:
    os.makedirs(os.path.dirname(p), exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        f.write(s)

def _log_name(base: str, tag: str) -> str:
    return os.path.join(LOGS_DIR, f"jaspergold_{base}{tag}.log")

def _log_base_from_tcl(tcl_path: str) -> str:
    base= os.path.basename(tcl_path)
    return os.path.splitext(base)[0]

def _decide_and_log(result: JGResult, tag: str):
    logger.info("[%s] status=%s,considered=%d,proven=%d,covered=%d",
                tag, result.status, result.properties_considered,result.assertions_proven, result.covers_covered)
    if result.raw_summary:
        logger.debug("SUMMARY BLOCK:\n%s", result.raw_summary)


def _find_modules(design_src: str) -> list[tuple[str, int, int]]:
    spans = []
    for m in re.finditer(r"\bmodule\s+([A-Za-z_]\w*)\b", design_src):
            name= m.group(1)
            start = m.start()
            mend= re.search(r"\bendmodule\b",design_src[m.end():], flags=re.S)
            end= m.end() + (mend.end() if mend else len(design_src))
            spans.append((name, start, end))
    return spans


def _choose_dut(mods: list[tuple[str, int, int]], property_name: str, category: str) -> tuple[str, int, int] | None:
    if not mods:
        return None

    pn = (property_name or "").lower()
    cat = (category or "").lower()
    for m in mods:
        if pn and pn in m[0].lower():
            return m

    family_hint = None
    if cat == "aes":
        family_hint = "aes"
    elif cat == "rsa":
        family_hint = "rsa"
    elif cat == "sha":
        family_hint = "sha"
    elif cat == "fsm":
        family_hint = "fsm"

    if family_hint:
        for m in mods:
            if family_hint in m[0].lower():
                return m

    return mods[0]
def _ensure_analysis(design_path: str, property_name: str, category: str, analysis_path: Optional[str]) -> str:
    if analysis_path and os.path.exists(analysis_path):
        return analysis_path

    design_src= _load_text(design_path)
    mods= _find_modules(design_src)
    chosen= _choose_dut(mods, property_name, category)
    detected =  ", ".join([m[0] for m in mods]) if mods else "(none found)"

    
    
    if chosen:
        name, s, e = chosen
        snippet= design_src[s:e]
        body= dedent(f"""
        Detected modules: {detected}

        Heuristic DUT selection for category "{category}" property "{property_name}": `{name}`

        ```verilog
        {snippet}
        ```
        """).strip()
    else:
        body = dedent(f"""
        Detected modules: {detected}

        No module block could be reliably extracted. 
        """).strip()

    out_path= os.path.join(GENERATED_DIR, f"{property_name}_analysis.txt")
    _write_text(out_path, body)
    return out_path

#Return---> {design_json_name:{header_func_name: best_module}}
def _read_summary_excel(summary_excel: str) -> Dict[str, Dict[str, str]]:
    try:
        import openpyxl
    except Exception as e:
        logger.warning("openpyxl not available   %s (%s)", summary_excel, e)
        return {}

    if not os.path.exists(summary_excel):
        logger.warning("Summary Excel not found: %s", summary_excel)
        return {}

    wb= openpyxl.load_workbook(summary_excel, data_only=True)
    ws= wb.active
    headers =  [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if not headers or headers[0].lower() != "design file":
        logger.warning("first header must be ''Design File'''): %s", summary_excel)
        return {}

    out: Dict[str, Dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        design_file = str(row[0]).strip()
        out[design_file] = {}
        for i in range(1, len(headers)):
                key = headers[i].strip()
                val = (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                out[design_file][key] = val
    return out


def _inject_best_module_into_analysis(
    analysis_path: str,
    family: str,
    function_name: str,
    design_sv_path: str,
    summary_excel: Optional[str]
) -> str:
    """
    Append:
      Best module for <FAMILY>:<function_name> → `module_name`
    """
    if not summary_excel:
        return analysis_path

    mapping = _read_summary_excel(summary_excel)
    if not mapping:
        return analysis_path

    sv_base = os.path.splitext(os.path.basename(design_sv_path))[0]
    json_name = sv_base + ".json"

    best = mapping.get(json_name, {}).get(function_name or "", "")
    if not best:
        logger.info("No summary entry for %s/%s in %s", json_name, function_name, summary_excel)
        return analysis_path

    text = _load_text(analysis_path)
    addon = f"\n\nBest module for {family}:{function_name} \u2192 `{best}`\n"
    _write_text(analysis_path, text + addon)
    logger.info("Injected best module from summary: %s", best)
    return analysis_path


###JasperGold runner 
def _run_local_jg(tcl_path: str, log_path: str) -> int:
    exe_jg = shutil.which("jg")
    exe_jaspergold = shutil.which("jaspergold")

    if exe_jg:
        cmd = [exe_jg, "-batch", os.path.abspath(tcl_path)]
    elif exe_jaspergold:
        cmd = [exe_jaspergold, "-nowindow", "-tcl", os.path.abspath(tcl_path)]
    else:
        _write_text(log_path, "ERROR:: neither `jg` nor `jaspergold` found on path.\n")
        return 127

    try:
        proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, check=False, text=True)
        _write_text(log_path, proc.stdout or "")
        return proc.returncode
    except Exception as e:
        _write_text(log_path, f"ERROR: failed to run {' '.join(cmd)}\n{e}\n")
        return 1



def run_jaspergold(tcl_path: str, log_path: str, design_path: str, wrapper_path: str) -> int:
    try:
        local_log, rc = run_verification(wrapper_path, tcl_path, design_path)
        try:
            if os.path.abspath(local_log) != os.path.abspath(log_path):
                shutil.copyfile(local_log, log_path)
        except Exception:
            _write_text(log_path, _load_text(local_log))
        return rc
    except Exception as e:
        logger.warning("Remote run_verification failed (%s). Falling back to local run.", e)
        return _run_local_jg(tcl_path, log_path)


def run_agentic_pipeline(
            category: str,
            design_path: Optional[str] = None,
            design_file: Optional[str] = None,
            property_name: Optional[str] = None,
            analysis_path: Optional[str] = None,
            design_label: Optional[str] = None,
            config: Optional[PipelineConfig] = None,
            llm_choice: Optional[str] = None,
            analysis_mode: str = "llm",
            generation_mode: str = "llm",
            **kwargs,
):
    if property_name is None:
        raise ValueError("property_name is required.")

    category = (category or "").strip()
    cat_upper = category.upper()

    design_path = design_path or design_file or kwargs.get("design")
    if not design_path:
        raise ValueError("design_path/design_file is required.")

    summary_excel = kwargs.get("summary_excel")

    fn_in = kwargs.get("function_name")
    default_fn_by_cat = {
            "RSA": "top_module",
            "SHA": "top_module",
            "FSM": "fsm_module",
    }
    function_name = fn_in or default_fn_by_cat.get(cat_upper, property_name)

    config = config or PipelineConfig()
    if analysis_path and os.path.exists(analysis_path):
        pass
    else:
        if analysis_mode == "llm":
            try:
                    props = load_properties(category)
                    prop_dict = next(
                        (p for p in props if str(p.get("name", "")).lower() == str(property_name).lower()),
                        {"name": property_name, "description": property_name},
                    )
            except Exception:
                prop_dict = {"name": property_name, "description": property_name}

            analysis_path = analyze_design(design_path, prop_dict, llm_choice=llm_choice)

            from tools.roles_table import  write_csv_json
            try:
                roles_json = write_csv_json(analysis_path, out_dir=GENERATED_DIR)
                logger.info(" wrote roles CSV/json from llm analysis → %s", roles_json)
            except Exception as e:
                logger.warning("could not extract roles table from analysis: %s", e)

        elif analysis_mode == "static":
            pdg_excel_path = kwargs.get("pdg_excel") or kwargs.get("pdg_excel_path")
            if not pdg_excel_path:
                raise ValueError("analysis_mode=static requires   --pdg_excel <path to .xlsx> (sheet=function_name).")

            import inspect
            #fn = static_bridge.static_analysis_from_excel_llm
            
            # params = set(inspect.signature(fn).parameters.keys())
            # excel_kw = None
            # for cand in ("pdg_excel", "pdg_excel_path", "excel_path"):
            #     if cand in params:
            #         excel_kw = cand
            #         break
            # if not excel_kw:
            #     raise TypeError(
            #         "error; "
            #         f"expected one of pdg_excel/pdg_excel_path/excel_path, got {sorted(params)}"
            #     )

            # sb_kwargs = {
            #         "family":  category,
            #         "function_name": function_name,
            #         "property_name":property_name,
            #         "llm_choice": llm_choice,
            #         excel_kw:pdg_excel_path,
            # }
            # analysis_path = fn(**sb_kwargs)

            # analysis_path = _inject_best_module_into_analysis(
            #     analysis_path, category, function_name, design_path, summary_excel
            # )

            from tools.roles_table import write_csv_json
            try:
                roles_json = write_csv_json(analysis_path, out_dir=GENERATED_DIR)
                logger.info("Wrote roles CSV/JSON from STATIC-LLM analysis → %s", roles_json)
            except Exception as e:
                logger.warning("Could not extract roles table from static analysis: %s", e)

        else:
            analysis_path = _ensure_analysis(design_path, property_name, category, None)

    ##generating wrapper + tcl
    if generation_mode == "algo":
        ##pname= re.sub( r"\W+", "", (property_name))
        pname = re.sub(r"\W+", "", (property_name or "")).lower()
        


        def _try_import(mod_name: str, fn_name: str):
            try:
                mod = importlib.import_module(f".{mod_name}", package=__package__)
                return getattr(mod, fn_name)
            except Exception as e:
                raise RuntimeError(
                    f"Algo dispatcher: could not load {mod_name}.{fn_name} → {e}"
                ) from e
        # def _try_import(mod_name: str, fn_name: str):
        #     try:
        #         mod= importlib.import_module(f".{mod_name}", package=__package__)
        #         return getattr(mod, fn_name)
        #     except Exception as e:
        #         logger.warning("Algo dispatcher: could not import %s.%s (%s). Falling back to generic.",
        #                        mod_name, fn_name, e)
        #         # from .sva_generator_algo import generate_sva_from_template as fallback_fn
        #         # return fallback_fn

        if cat_upper == "AES":
            aes_dispatch = {
                "addroundkey":   ("sva_generator_alg_ip_AES_AddRoundKey","generate_sva_for_aes_addroundkey"),
                "sbox":         ("sva_generator_alg_ip_AES_SBox","generate_sva_for_aes_property"),
                "shiftrows":    ("sva_generator_alg_ip_AES_ShiftRows", "generate_sva_for_aes_shiftrows"),
                "keyexpansion": ("sva_generator_alg_ip_AES_KeyExp", "generate_sva_for_aes_keyexp"),
            }
            mod_name, fn_name=aes_dispatch.get(pname, ("sva_generator_algo","generate_sva_from_template"))
            gen_fn = _try_import(mod_name, fn_name)

        elif cat_upper == "FSM":
            fsm_dispatch = {
                 "always_legal_state": ("sva_generator_alg_ip_FSM", "generate_sva_for_fsm_property"),
                 "recovery_from_illegal_state": ("sva_generator_alg_ip_FSM","generate_sva_for_fsm_property"),
            }
            mod_name, fn_name = fsm_dispatch.get(pname, ("sva_generator_algo","generate_sva_from_template"))
            gen_fn= _try_import(mod_name, fn_name)

        elif cat_upper == "RSA":
              gen_fn= _try_import("sva_generator_alg_ip_RSA", "generate_sva_for_rsa_property")

        elif cat_upper == "SHA":
            gen_fn = _try_import("sva_generator_alg_ip_SHA", "generate_sva_for_sha_property")

        # else:
        #     from .sva_generator_algo import generate_sva_from_template as gen_fn

        wrapper_sv_path, tcl_path, top = gen_fn(
            category=category,
            property_name=property_name,
            design_path=design_path,
            analysis_path=analysis_path,
        )

    else:
        wrapper_sv_path, tcl_path, top = generate_sva(
                category=category,
                property_name=property_name,
                design_path=design_path,
                analysis_path=analysis_path,
                design_label=design_label,
                llm_choice=llm_choice,
        )

    log_base = _log_base_from_tcl(tcl_path)
    log_path = _log_name(log_base, "")
    run_jaspergold(tcl_path, log_path, design_path, wrapper_sv_path)
    log_text = _load_text(log_path)

    ###analysissss
    res = analyze_jg_log(log_text)
    _decide_and_log(res, f"{property_name}")

    if res.status== "pass":
        return   {"result": "Pass", "wrapper": wrapper_sv_path, "tcl": tcl_path,"top": top,
                "counts": {"considered":res.properties_considered, "proven": res.assertions_proven,
                           "covered":res.covers_covered, "errors":res.errors}}
    if res.status == "fail":
        return {"result": "Fail","wrapper":wrapper_sv_path, "tcl":tcl_path, "top":top,
                "counts": {"considered": res.properties_considered,"proven":res.assertions_proven,
                           "covered":res.covers_covered,"errors": res.errors},
                "issues": res.issues}
    if generation_mode == "algo":
        final = "NoProperties"   if res.status == "no_properties"   else "Error"
        return {"result": final,"last_status": res.status, "attempts": 0, "issues": res.issues}

    
    
    ##repair loop(for LLM gen only)
    attempts=0
    fixed_sv= wrapper_sv_path
    fixed_tcl = tcl_path
    while attempts < config.max_repairs and res.status in ("error", "no_properties"):
        attempts += 1
        logger.info("Attempting repair %d ...", attempts)

        fixed_sv, fixed_tcl, fixed_top = repair_wrapper(
                    property_name=property_name,
                    failing_wrapper_sv_path=fixed_sv,
                    failing_tcl_path=fixed_tcl,
                    jg_log_text=log_text,
                    analysis_text=_load_text(analysis_path),
                    attempt_idx=attempts,
                    design_src_path=design_path,
                    llm_choice=llm_choice,
        )

        rbase = _log_base_from_tcl(fixed_tcl)
        
        rlog_path = _log_name(rbase, f"")
        run_jaspergold(fixed_tcl, rlog_path, design_path, fixed_sv)
        log_text = _load_text(rlog_path)

        res = analyze_jg_log(log_text)
        _decide_and_log(res, f"{property_name} (repair {attempts})")

        if res.status == "pass":
            return  {"result": "Pass","wrapper": fixed_sv,"tcl": fixed_tcl, "top": fixed_top}
        if res.status == "fail":
            return     {"result": "Fail","wrapper": fixed_sv, "tcl": fixed_tcl,"top": fixed_top,
                    "counts": {"considered": res.properties_considered, "proven": res.assertions_proven,
                               "covered": res.covers_covered, "errors": res.errors},
                    "issues": res.issues}

    final= "NoProperties" if res.status == "no_properties" else "Error"
    return {"result":final, "last_status":res.status, "attempts":attempts, "issues":res.issues}

